import os
import re
import math
import hashlib
from datetime import date, datetime, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import text, bindparam
from openpyxl import load_workbook

SECRET_SALT = os.getenv("SECRET_SALT")
if not SECRET_SALT:
    raise RuntimeError("SECRET_SALT is not set")

SLOT_DAY = "DAY"
SLOT_FULL_INVENT = "FULL_INVENT"

DEFAULT_RATE_SUPPLY = 800
DEFAULT_RATE_NO_SUPPLY = 400
DEFAULT_RATE_INVENTORY = 400
DEFAULT_RATE_COFFEE = 100


def fio_norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("ё", "е")
    s = re.sub(r"[\u00A0\u2000-\u200B\u202F\u205F\u3000]", " ", s)
    s = re.sub(r"[^а-яa-z\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def hash_last4(last4: str) -> str:
    s = (last4.strip() + SECRET_SALT).encode("utf-8")
    return hashlib.sha256(s).hexdigest()


def get_active_period():
    today = date.today()
    if today.day <= 5:
        if today.month == 1:
            report_year = today.year - 1
            report_month = 12
        else:
            report_year = today.year
            report_month = today.month - 1
    else:
        report_year = today.year
        report_month = today.month

    editable_until_year = report_year
    editable_until_month = report_month + 1
    if editable_until_month == 13:
        editable_until_month = 1
        editable_until_year += 1

    editable_until = date(editable_until_year, editable_until_month, 5)
    return {
        "year": report_year,
        "month": report_month,
        "editable_until": editable_until.isoformat(),
    }


def get_merchants_columns(db: Session):
    query = text(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_name = 'merchants'
        ORDER BY ordinal_position
        """
    )
    rows = db.execute(query).fetchall()
    return [row[0] for row in rows]


def login_user(db: Session, fio: str, last4: str):
    fio_n = fio_norm(fio)
    result = db.execute(
        text(
            """
            SELECT id, fio, fio_norm, pass_hash, telegram_id, tu, created_at
            FROM merchants
            WHERE fio_norm = :fio_norm
            LIMIT 1
            """
        ),
        {"fio_norm": fio_n},
    ).mappings().first()

    if not result:
        return None
    if hash_last4(last4) != result["pass_hash"]:
        return None

    return {
        "id": result["id"],
        "fio": result["fio"],
        "fio_norm": result["fio_norm"],
        "telegram_id": result["telegram_id"],
        "tu": result["tu"],
        "created_at": str(result["created_at"]) if result["created_at"] else None,
    }


def get_merchant_by_fio(db: Session, fio: str):
    fio_n = fio_norm(fio)
    result = db.execute(
        text(
            """
            SELECT id, fio, fio_norm, telegram_id, tu, created_at
            FROM merchants
            WHERE fio_norm = :fio_norm
            LIMIT 1
            """
        ),
        {"fio_norm": fio_n},
    ).mappings().first()

    if not result:
        return None

    return {
        "id": result["id"],
        "fio": result["fio"],
        "fio_norm": result["fio_norm"],
        "telegram_id": result["telegram_id"],
        "tu": result["tu"],
        "created_at": str(result["created_at"]) if result["created_at"] else None,
    }


def month_start(y: int, m: int) -> date:
    return date(y, m, 1)


def month_end_exclusive(y: int, m: int) -> date:
    return date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)


def days_in_month(y: int, m: int) -> int:
    return (month_end_exclusive(y, m) - timedelta(days=1)).day


def weekday_of(y: int, m: int, d: int) -> int:
    return date(y, m, d).weekday()


def ensure_special_inventory_days_table(db: Session):
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS special_inventory_days (
            id SERIAL PRIMARY KEY,
            inv_date DATE UNIQUE NOT NULL
        )
    """))
    db.commit()


def add_special_inventory_day(db: Session, inv_date: date):
    ensure_special_inventory_days_table(db)
    db.execute(text("""
        INSERT INTO special_inventory_days (inv_date)
        VALUES (:inv_date)
        ON CONFLICT (inv_date) DO NOTHING
    """), {"inv_date": inv_date})
    db.commit()


def delete_special_inventory_day(db: Session, inv_date: date):
    ensure_special_inventory_days_table(db)
    db.execute(text("DELETE FROM special_inventory_days WHERE inv_date = :inv_date"), {"inv_date": inv_date})
    db.commit()


def get_special_inventory_days(db: Session) -> list[date]:
    ensure_special_inventory_days_table(db)
    rows = db.execute(text("SELECT inv_date FROM special_inventory_days ORDER BY inv_date")).all()
    return [r[0] for r in rows if r and r[0]]


def get_special_inventory_days_set(db: Session) -> set[date]:
    return set(get_special_inventory_days(db))


def is_inventory_allowed_date(db: Session, current_date: date) -> bool:
    return current_date.weekday() in (4, 5) or current_date in get_special_inventory_days_set(db)


def month_title(y: int, m: int) -> str:
    names = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
    ]
    return f"{names[m - 1]} {y}"


def normalize_point_code(v) -> str:
    s = str(v or "").strip()
    s = re.sub(r"\s+", "", s)
    return s


def point_has_any_supply_in_month(db: Session, point_code: str, y: int, m: int) -> bool:
    start = month_start(y, m)
    end = month_end_exclusive(y, m)
    result = db.execute(
        text(
            """
            SELECT 1
            FROM supplies
            WHERE point_code = :point_code
              AND supply_date >= :start_date
              AND supply_date < :end_date
            LIMIT 1
            """
        ),
        {"point_code": point_code, "start_date": start, "end_date": end},
    ).first()
    return result is not None


def get_supply_boxes_map(db: Session, point_code: str, y: int, m: int) -> dict[int, int]:
    start = month_start(y, m)
    end = month_end_exclusive(y, m)
    rows = db.execute(
        text(
            """
            SELECT supply_date, boxes
            FROM supplies
            WHERE point_code = :point_code
              AND supply_date >= :start_date
              AND supply_date < :end_date
            ORDER BY supply_date
            """
        ),
        {"point_code": point_code, "start_date": start, "end_date": end},
    ).mappings().all()
    result: dict[int, int] = {}
    for row in rows:
        result[row["supply_date"].day] = int(row["boxes"] or 0)
    return result



def get_supply_days_for_point(db: Session, point_code: str, y: int, m: int) -> list[date]:
    start = month_start(y, m)
    end = month_end_exclusive(y, m)
    rows = db.execute(text("""
        SELECT supply_date
        FROM supplies
        WHERE point_code = :point_code
          AND supply_date >= :start_date
          AND supply_date < :end_date
          AND COALESCE(boxes, 0) > 0
        ORDER BY supply_date
    """), {
        "point_code": point_code,
        "start_date": start,
        "end_date": end,
    }).all()
    return [r[0] for r in rows if r and r[0]]


def get_supply_adjustment_amount(db: Session, point_code: str, y: int, m: int) -> int:
    rates = get_point_rates(db, point_code, y, m)
    diff = int(rates["rate_supply"] or 0) - int(rates["rate_no_supply"] or 0)
    return -max(0, diff)


def get_visits_for_month(db: Session, merchant_id: int, point_code: str, y: int, m: int) -> dict[int, set[str]]:
    start = month_start(y, m)
    end = month_end_exclusive(y, m)
    rows = db.execute(
        text(
            """
            SELECT visit_date, slot
            FROM visits
            WHERE merchant_id = :merchant_id
              AND point_code = :point_code
              AND visit_date >= :start_date
              AND visit_date < :end_date
            """
        ),
        {
            "merchant_id": merchant_id,
            "point_code": point_code,
            "start_date": start,
            "end_date": end,
        },
    ).mappings().all()

    result: dict[int, set[str]] = {}
    for row in rows:
        day = row["visit_date"].day
        result.setdefault(day, set()).add(str(row["slot"]))
    return result


def toggle_day_visit(db: Session, merchant_id: int, point_code: str, y: int, m: int, day: int):
    visit_date = date(y, m, day)
    existing = db.execute(
        text(
            """
            SELECT id
            FROM visits
            WHERE merchant_id = :merchant_id
              AND point_code = :point_code
              AND visit_date = :visit_date
              AND slot = :slot
            LIMIT 1
            """
        ),
        {
            "merchant_id": merchant_id,
            "point_code": point_code,
            "visit_date": visit_date,
            "slot": SLOT_DAY,
        },
    ).scalar()

    if existing:
        db.execute(text("DELETE FROM visits WHERE id = :id"), {"id": existing})
        db.commit()
        return "removed"

    db.execute(
        text(
            """
            INSERT INTO visits (merchant_id, point_code, visit_date, slot)
            VALUES (:merchant_id, :point_code, :visit_date, :slot)
            ON CONFLICT DO NOTHING
            """
        ),
        {
            "merchant_id": merchant_id,
            "point_code": point_code,
            "visit_date": visit_date,
            "slot": SLOT_DAY,
        },
    )
    db.commit()
    return "added"


def toggle_inventory_visit(db: Session, merchant_id: int, point_code: str, y: int, m: int, day: int):
    visit_date = date(y, m, day)
    existing = db.execute(
        text(
            """
            SELECT id
            FROM visits
            WHERE merchant_id = :merchant_id
              AND point_code = :point_code
              AND visit_date = :visit_date
              AND slot = :slot
            LIMIT 1
            """
        ),
        {
            "merchant_id": merchant_id,
            "point_code": point_code,
            "visit_date": visit_date,
            "slot": SLOT_FULL_INVENT,
        },
    ).scalar()

    if existing:
        db.execute(text("DELETE FROM visits WHERE id = :id"), {"id": existing})
        db.commit()
        return "removed"

    db.execute(
        text(
            """
            INSERT INTO visits (merchant_id, point_code, visit_date, slot)
            VALUES (:merchant_id, :point_code, :visit_date, :slot)
            ON CONFLICT DO NOTHING
            """
        ),
        {
            "merchant_id": merchant_id,
            "point_code": point_code,
            "visit_date": visit_date,
            "slot": SLOT_FULL_INVENT,
        },
    )
    db.commit()
    return "added"


def get_point_rates(db: Session, point_code: str, y: int, m: int):
    mk = month_start(y, m)
    row = db.execute(
        text(
            """
            SELECT rate_supply, rate_no_supply, rate_inventory, coffee_enabled, coffee_rate, pay_lt5
            FROM point_rates
            WHERE point_code = :point_code
              AND month_key = :month_key
            LIMIT 1
            """
        ),
        {"point_code": point_code, "month_key": mk},
    ).mappings().first()

    if not row:
        return {
            "rate_supply": DEFAULT_RATE_SUPPLY,
            "rate_no_supply": DEFAULT_RATE_NO_SUPPLY,
            "rate_inventory": DEFAULT_RATE_INVENTORY,
            "coffee_enabled": False,
            "coffee_rate": DEFAULT_RATE_COFFEE,
            "pay_lt5": False,
        }

    return {
        "rate_supply": int(row["rate_supply"] or DEFAULT_RATE_SUPPLY),
        "rate_no_supply": int(row["rate_no_supply"] or DEFAULT_RATE_NO_SUPPLY),
        "rate_inventory": int(row["rate_inventory"] or DEFAULT_RATE_INVENTORY),
        "coffee_enabled": bool(row["coffee_enabled"]),
        "coffee_rate": int(row["coffee_rate"] or DEFAULT_RATE_COFFEE),
        "pay_lt5": bool(row["pay_lt5"]),
    }


def effective_has_supply(boxes: int, pay_lt5: bool) -> bool:
    if boxes <= 0:
        return False
    return True if pay_lt5 else (boxes >= 5)


def ensure_monthly_submissions_table(db: Session):
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS monthly_submissions (
                id SERIAL PRIMARY KEY,
                merchant_id INTEGER NOT NULL,
                month_key DATE NOT NULL,
                comment TEXT,
                extra_amount INTEGER NOT NULL DEFAULT 0,
                receipt_path TEXT,
                status TEXT NOT NULL DEFAULT 'draft',
                created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMP NOT NULL DEFAULT NOW(),
                UNIQUE (merchant_id, month_key)
            )
            """
        )
    )
    db.commit()


def get_monthly_submission(db: Session, merchant_id: int, y: int, m: int):
    ensure_monthly_submissions_table(db)
    mk = month_start(y, m)
    row = db.execute(
        text(
            """
            SELECT *
            FROM monthly_submissions
            WHERE merchant_id = :merchant_id
              AND month_key = :month_key
            LIMIT 1
            """
        ),
        {"merchant_id": merchant_id, "month_key": mk},
    ).mappings().first()
    return dict(row) if row else None


def upsert_monthly_submission_draft(db: Session, merchant_id: int, y: int, m: int, comment: str, extra_amount: int, receipt_path: str | None):
    ensure_monthly_submissions_table(db)
    mk = month_start(y, m)
    existing = get_monthly_submission(db, merchant_id, y, m)

    if existing:
        if receipt_path:
            db.execute(text("""
                UPDATE monthly_submissions
                SET comment=:comment, extra_amount=:extra_amount, receipt_path=:receipt_path, updated_at=NOW()
                WHERE merchant_id=:merchant_id AND month_key=:month_key
            """), {
                "comment": comment,
                "extra_amount": extra_amount,
                "receipt_path": receipt_path,
                "merchant_id": merchant_id,
                "month_key": mk,
            })
        else:
            db.execute(text("""
                UPDATE monthly_submissions
                SET comment=:comment, extra_amount=:extra_amount, updated_at=NOW()
                WHERE merchant_id=:merchant_id AND month_key=:month_key
            """), {
                "comment": comment,
                "extra_amount": extra_amount,
                "merchant_id": merchant_id,
                "month_key": mk,
            })
    else:
        db.execute(text("""
            INSERT INTO monthly_submissions (merchant_id, month_key, comment, extra_amount, receipt_path, status)
            VALUES (:merchant_id, :month_key, :comment, :extra_amount, :receipt_path, 'draft')
        """), {
            "merchant_id": merchant_id,
            "month_key": mk,
            "comment": comment,
            "extra_amount": extra_amount,
            "receipt_path": receipt_path,
        })
    db.commit()


def submit_monthly_submission(db: Session, merchant_id: int, y: int, m: int):
    ensure_monthly_submissions_table(db)
    mk = month_start(y, m)
    db.execute(text("""
        INSERT INTO monthly_submissions (merchant_id, month_key, comment, extra_amount, receipt_path, status)
        VALUES (:merchant_id, :month_key, '', 0, NULL, 'submitted')
        ON CONFLICT (merchant_id, month_key)
        DO UPDATE SET status='submitted', updated_at=NOW()
    """), {"merchant_id": merchant_id, "month_key": mk})
    db.commit()


def reopen_monthly_submission(db: Session, merchant_id: int, y: int, m: int):
    ensure_monthly_submissions_table(db)
    mk = month_start(y, m)
    db.execute(text("""
        UPDATE monthly_submissions
        SET status='draft', updated_at=NOW()
        WHERE merchant_id=:merchant_id AND month_key=:month_key
    """), {"merchant_id": merchant_id, "month_key": mk})
    db.commit()


def get_points_for_month(db: Session, merchant_id: int, y: int, m: int) -> list[str]:
    start = month_start(y, m)
    end = month_end_exclusive(y, m)
    rows = db.execute(text("""
        SELECT DISTINCT point_code
        FROM visits
        WHERE merchant_id=:merchant_id
          AND visit_date >= :start_date
          AND visit_date < :end_date
        ORDER BY point_code
    """), {"merchant_id": merchant_id, "start_date": start, "end_date": end}).all()
    return [r[0] for r in rows if r and r[0]]


def ensure_point_adjustments_table(db: Session):
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS point_adjustments (
            id SERIAL PRIMARY KEY,
            merchant_id INTEGER NOT NULL,
            point_code TEXT NOT NULL,
            month_key DATE NOT NULL,
            note_amount INTEGER NOT NULL DEFAULT 0,
            note_comment TEXT,
            reimb_amount INTEGER NOT NULL DEFAULT 0,
            reimb_comment TEXT,
            reimb_receipt TEXT,
            created_at TIMESTAMP NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMP NOT NULL DEFAULT NOW(),
            UNIQUE (merchant_id, point_code, month_key)
        )
    """))
    db.commit()


def get_point_adjustment(db: Session, merchant_id: int, point_code: str, y: int, m: int):
    ensure_point_adjustments_table(db)
    mk = month_start(y, m)
    row = db.execute(text("""
        SELECT *
        FROM point_adjustments
        WHERE merchant_id = :merchant_id
          AND point_code = :point_code
          AND month_key = :month_key
        LIMIT 1
    """), {
        "merchant_id": merchant_id,
        "point_code": point_code,
        "month_key": mk
    }).mappings().first()
    return dict(row) if row else None


def upsert_point_adjustment(
    db: Session,
    merchant_id: int,
    point_code: str,
    y: int,
    m: int,
    note_amount: int,
    note_comment: str,
    reimb_amount: int,
    reimb_comment: str,
    reimb_receipt: str | None,
):
    ensure_point_adjustments_table(db)
    mk = month_start(y, m)
    existing = get_point_adjustment(db, merchant_id, point_code, y, m)

    if existing:
        db.execute(text("""
            UPDATE point_adjustments
            SET note_amount = :note_amount,
                note_comment = :note_comment,
                reimb_amount = :reimb_amount,
                reimb_comment = :reimb_comment,
                reimb_receipt = COALESCE(:reimb_receipt, reimb_receipt),
                updated_at = NOW()
            WHERE merchant_id = :merchant_id
              AND point_code = :point_code
              AND month_key = :month_key
        """), {
            "merchant_id": merchant_id,
            "point_code": point_code,
            "month_key": mk,
            "note_amount": note_amount,
            "note_comment": note_comment,
            "reimb_amount": reimb_amount,
            "reimb_comment": reimb_comment,
            "reimb_receipt": reimb_receipt,
        })
    else:
        db.execute(text("""
            INSERT INTO point_adjustments (
                merchant_id, point_code, month_key,
                note_amount, note_comment, reimb_amount, reimb_comment, reimb_receipt
            ) VALUES (
                :merchant_id, :point_code, :month_key,
                :note_amount, :note_comment, :reimb_amount, :reimb_comment, :reimb_receipt
            )
        """), {
            "merchant_id": merchant_id,
            "point_code": point_code,
            "month_key": mk,
            "note_amount": note_amount,
            "note_comment": note_comment,
            "reimb_amount": reimb_amount,
            "reimb_comment": reimb_comment,
            "reimb_receipt": reimb_receipt,
        })

    db.commit()

def compute_point_total(db: Session, merchant_id: int, point_code: str, y: int, m: int):
    ensure_point_adjustments_table(db)
    boxes_map = get_supply_boxes_map(db, point_code, y, m)
    visits = get_visits_for_month(db, merchant_id, point_code, y, m)
    rates = get_point_rates(db, point_code, y, m)
    point_adj = get_point_adjustment(db, merchant_id, point_code, y, m)

    total = 0
    cnt_supply = 0
    cnt_no_supply = 0
    cnt_day_total = 0
    cnt_full_inv = 0
    sum_supply = 0
    sum_no_supply = 0
    sum_inventory = 0

    for day, slots in visits.items():
        if SLOT_DAY in slots:
            cnt_day_total += 1
            boxes = boxes_map.get(day, 0)
            if effective_has_supply(boxes, rates["pay_lt5"]):
                cnt_supply += 1
                total += rates["rate_supply"]
                sum_supply += rates["rate_supply"]
            else:
                cnt_no_supply += 1
                total += rates["rate_no_supply"]
                sum_no_supply += rates["rate_no_supply"]

        if SLOT_FULL_INVENT in slots:
            cnt_full_inv += 1
            total += rates["rate_inventory"]
            sum_inventory += rates["rate_inventory"]

    coffee_sum = 0
    coffee_cnt = 0
    if rates["coffee_enabled"] and cnt_day_total > 0:
        coffee_cnt = cnt_day_total
        coffee_sum = rates["coffee_rate"] * cnt_day_total
        total += coffee_sum

    note_amount = 0
    note_comment = ""
    reimb_amount = 0
    reimb_comment = ""
    reimb_receipt = None

    if point_adj:
        note_amount = int(point_adj.get("note_amount") or 0)
        note_comment = point_adj.get("note_comment") or ""
        reimb_amount = int(point_adj.get("reimb_amount") or 0)
        reimb_comment = point_adj.get("reimb_comment") or ""
        reimb_receipt = point_adj.get("reimb_receipt")
        total += note_amount + reimb_amount

    return {
        "total": total,
        "cnt_supply": cnt_supply,
        "cnt_no_supply": cnt_no_supply,
        "cnt_day_total": cnt_day_total,
        "cnt_full_inv": cnt_full_inv,
        "sum_supply": sum_supply,
        "sum_no_supply": sum_no_supply,
        "sum_inventory": sum_inventory,
        "coffee_enabled": rates["coffee_enabled"],
        "coffee_rate": rates["coffee_rate"],
        "coffee_sum": coffee_sum,
        "coffee_cnt": coffee_cnt,
        "pay_lt5": rates["pay_lt5"],
        "rate_supply": rates["rate_supply"],
        "rate_no_supply": rates["rate_no_supply"],
        "rate_inventory": rates["rate_inventory"],
        "note_amount": note_amount,
        "note_comment": note_comment,
        "reimb_amount": reimb_amount,
        "reimb_comment": reimb_comment,
        "reimb_receipt": reimb_receipt,
    }


def compute_overall_total(db: Session, merchant_id: int, y: int, m: int):
    ensure_monthly_submissions_table(db)
    points = get_points_for_month(db, merchant_id, y, m)
    total = 0
    per_point = {}
    per_point_details = {}

    for point_code in points:
        detail = compute_point_total(db, merchant_id, point_code, y, m)
        per_point[point_code] = detail["total"]
        per_point_details[point_code] = detail
        total += detail["total"]

    monthly = get_monthly_submission(db, merchant_id, y, m)
    extra_amount = 0
    comment = ""
    receipt_path = None
    submission_status = "draft"
    if monthly:
        extra_amount = int(monthly.get("extra_amount") or 0)
        comment = monthly.get("comment") or ""
        receipt_path = monthly.get("receipt_path")
        submission_status = monthly.get("status") or "draft"
        total += extra_amount

    return {
        "total": total,
        "per_point": per_point,
        "per_point_details": per_point_details,
        "extra_amount": extra_amount,
        "comment": comment,
        "receipt_path": receipt_path,
        "submission_status": submission_status,
    }


def get_all_tu_values(db: Session) -> list[str]:
    rows = db.execute(text("""
        SELECT DISTINCT tu
        FROM merchants
        WHERE tu IS NOT NULL AND TRIM(tu) <> ''
        ORDER BY tu
    """)).all()
    return [r[0] for r in rows if r and r[0]]


def get_admin_report_rows(db: Session, y: int, m: int, tu: str | None = None, status: str | None = None):
    """Fast admin report builder: one grouped SQL query instead of per-row recalculation."""
    ensure_monthly_submissions_table(db)
    ensure_point_adjustments_table(db)
    start = month_start(y, m)
    end = month_end_exclusive(y, m)

    params = {
        "start_date": start,
        "end_date": end,
        "month_key": start,
        "slot_day": SLOT_DAY,
        "slot_full_invent": SLOT_FULL_INVENT,
        "default_rate_supply": DEFAULT_RATE_SUPPLY,
        "default_rate_no_supply": DEFAULT_RATE_NO_SUPPLY,
        "default_rate_inventory": DEFAULT_RATE_INVENTORY,
        "default_rate_coffee": DEFAULT_RATE_COFFEE,
    }

    tu_sql = ""
    if tu:
        tu_sql = " AND m.tu = :tu"
        params["tu"] = tu

    sql = f"""
        WITH overlap_rows AS (
            SELECT DISTINCT v1.merchant_id, v1.point_code
            FROM visits v1
            JOIN visits v2
              ON v1.visit_date = v2.visit_date
             AND v1.point_code = v2.point_code
             AND v1.slot = v2.slot
             AND v1.merchant_id <> v2.merchant_id
            WHERE v1.visit_date >= :start_date
              AND v1.visit_date < :end_date
        )
        SELECT
            v.merchant_id,
            v.point_code,
            m.fio,
            COALESCE(m.tu, '') AS tu,
            COALESCE(ms.status, 'не отправлено') AS status,
            COALESCE(ms.comment, '') AS monthly_comment,
            COALESCE(ms.extra_amount, 0) AS monthly_extra_amount,
            ms.receipt_path AS monthly_receipt_path,

            COALESCE(pr.rate_supply, :default_rate_supply) AS rate_supply,
            COALESCE(pr.rate_no_supply, :default_rate_no_supply) AS rate_no_supply,
            COALESCE(pr.rate_inventory, :default_rate_inventory) AS rate_inventory,
            COALESCE(pr.coffee_enabled, FALSE) AS coffee_enabled,
            COALESCE(pr.coffee_rate, :default_rate_coffee) AS coffee_rate,
            COALESCE(pr.pay_lt5, FALSE) AS pay_lt5,

            COALESCE(pa.note_amount, 0) AS note_amount,
            COALESCE(pa.note_comment, '') AS note_comment,
            COALESCE(pa.reimb_amount, 0) AS reimb_amount,
            COALESCE(pa.reimb_comment, '') AS reimb_comment,
            pa.reimb_receipt AS reimb_receipt,
            CASE WHEN ov.merchant_id IS NULL THEN FALSE ELSE TRUE END AS has_overlap,

            SUM(CASE
                WHEN v.slot = :slot_day
                 AND COALESCE(s.boxes, 0) > 0
                 AND (COALESCE(pr.pay_lt5, FALSE) = TRUE OR COALESCE(s.boxes, 0) >= 5)
                THEN 1 ELSE 0 END) AS cnt_supply,

            SUM(CASE
                WHEN v.slot = :slot_day
                 AND NOT (
                    COALESCE(s.boxes, 0) > 0
                    AND (COALESCE(pr.pay_lt5, FALSE) = TRUE OR COALESCE(s.boxes, 0) >= 5)
                 )
                THEN 1 ELSE 0 END) AS cnt_no_supply,

            SUM(CASE WHEN v.slot = :slot_full_invent THEN 1 ELSE 0 END) AS cnt_full_inv,
            SUM(CASE WHEN v.slot = :slot_day THEN 1 ELSE 0 END) AS cnt_day_total
        FROM visits v
        JOIN merchants m ON m.id = v.merchant_id
        LEFT JOIN supplies s
          ON s.point_code = v.point_code
         AND s.supply_date = v.visit_date
        LEFT JOIN point_rates pr
          ON pr.point_code = v.point_code
         AND pr.month_key = :month_key
        LEFT JOIN point_adjustments pa
          ON pa.merchant_id = v.merchant_id
         AND pa.point_code = v.point_code
         AND pa.month_key = :month_key
        LEFT JOIN monthly_submissions ms
          ON ms.merchant_id = v.merchant_id
         AND ms.month_key = :month_key
        LEFT JOIN overlap_rows ov
          ON ov.merchant_id = v.merchant_id
         AND ov.point_code = v.point_code
        WHERE v.visit_date >= :start_date
          AND v.visit_date < :end_date
          {tu_sql}
        GROUP BY
            v.merchant_id, v.point_code, m.fio, m.tu,
            ms.status, ms.comment, ms.extra_amount, ms.receipt_path,
            pr.rate_supply, pr.rate_no_supply, pr.rate_inventory,
            pr.coffee_enabled, pr.coffee_rate, pr.pay_lt5,
            pa.note_amount, pa.note_comment, pa.reimb_amount, pa.reimb_comment, pa.reimb_receipt,
            ov.merchant_id
        ORDER BY m.tu NULLS LAST, m.fio, v.point_code
    """

    rows = db.execute(text(sql), params).mappings().all()
    result = []
    for row in rows:
        status_value = row["status"] or "не отправлено"
        if status and status_value != status:
            continue

        cnt_supply = int(row["cnt_supply"] or 0)
        cnt_no_supply = int(row["cnt_no_supply"] or 0)
        cnt_full_inv = int(row["cnt_full_inv"] or 0)
        cnt_day_total = int(row["cnt_day_total"] or 0)
        rate_supply = int(row["rate_supply"] or DEFAULT_RATE_SUPPLY)
        rate_no_supply = int(row["rate_no_supply"] or DEFAULT_RATE_NO_SUPPLY)
        rate_inventory = int(row["rate_inventory"] or DEFAULT_RATE_INVENTORY)
        coffee_rate = int(row["coffee_rate"] or DEFAULT_RATE_COFFEE)
        coffee_enabled = bool(row["coffee_enabled"])
        note_amount = int(row["note_amount"] or 0)
        reimb_amount = int(row["reimb_amount"] or 0)

        sum_supply = cnt_supply * rate_supply
        sum_no_supply = cnt_no_supply * rate_no_supply
        sum_inventory = cnt_full_inv * rate_inventory
        coffee_cnt = cnt_day_total if coffee_enabled else 0
        coffee_sum = coffee_cnt * coffee_rate if coffee_enabled else 0
        point_total = sum_supply + sum_no_supply + sum_inventory + coffee_sum + note_amount + reimb_amount

        result.append({
            "merchant_id": row["merchant_id"],
            "fio": row["fio"],
            "tu": row["tu"] or "",
            "point_code": row["point_code"],
            "month_key": str(start),
            "status": status_value,
            "comment": row["monthly_comment"] or "",
            "extra_amount": int(row["monthly_extra_amount"] or 0),
            "receipt_path": row["monthly_receipt_path"],
            "note_amount": note_amount,
            "note_comment": row["note_comment"] or "",
            "reimb_amount": reimb_amount,
            "reimb_comment": row["reimb_comment"] or "",
            "reimb_receipt": row["reimb_receipt"],
            "cnt_supply": cnt_supply,
            "cnt_no_supply": cnt_no_supply,
            "cnt_total_exits": cnt_supply + cnt_no_supply,
            "cnt_full_inv": cnt_full_inv,
            "sum_supply": sum_supply,
            "sum_no_supply": sum_no_supply,
            "sum_inventory": sum_inventory,
            "coffee_enabled": coffee_enabled,
            "coffee_cnt": coffee_cnt,
            "coffee_rate": coffee_rate,
            "coffee_sum": coffee_sum,
            "has_overlap": bool(row["has_overlap"]),
            "point_total": point_total,
        })
    return result


def get_admin_payroll_rows(db: Session, y: int, m: int, tu: str | None = None, status: str | None = None):
    rows = get_admin_report_rows(db, y, m, tu, status)
    grouped: dict[int, dict] = {}
    for r in rows:
        mid = r["merchant_id"]
        if mid not in grouped:
            grouped[mid] = {
                "merchant_id": mid,
                "fio": r["fio"],
                "tu": r["tu"] or "",
                "clean_total": 0,
                "status": r["status"],
            }
        grouped[mid]["clean_total"] += int(r["point_total"] or 0)
        if grouped[mid]["status"] != "submitted" and r["status"] == "submitted":
            grouped[mid]["status"] = "submitted"

    result = []
    for row in grouped.values():
        clean_total = int(row["clean_total"] or 0)
        row["payroll_total"] = math.ceil(clean_total / 0.87) if clean_total > 0 else 0
        result.append(row)
    result.sort(key=lambda x: (x.get("tu") or "", x.get("fio") or ""))
    return result


def get_intersections_rows(db: Session, y: int, m: int, tu: str | None = None):
    start = month_start(y, m)
    end = month_end_exclusive(y, m)
    sql = """
        SELECT v1.visit_date, v1.point_code,
               m1.fio AS fio1, m1.tu AS tu1,
               m2.fio AS fio2, m2.tu AS tu2,
               v1.slot AS slot1, v2.slot AS slot2
        FROM visits v1
        JOIN visits v2
          ON v1.visit_date = v2.visit_date
         AND v1.point_code = v2.point_code
         AND v1.slot = v2.slot
         AND v1.merchant_id < v2.merchant_id
        JOIN merchants m1 ON m1.id = v1.merchant_id
        JOIN merchants m2 ON m2.id = v2.merchant_id
        WHERE v1.visit_date >= :start_date
          AND v1.visit_date < :end_date
    """
    params = {"start_date": start, "end_date": end}
    if tu:
        sql += " AND (m1.tu = :tu OR m2.tu = :tu)"
        params["tu"] = tu
    sql += " ORDER BY v1.visit_date, v1.point_code, v1.slot, m1.fio, m2.fio"
    rows = db.execute(text(sql), params).mappings().all()
    return [{
        "visit_date": str(r["visit_date"]),
        "point_code": r["point_code"],
        "fio1": r["fio1"],
        "tu1": r["tu1"] or "",
        "slot1": r["slot1"],
        "fio2": r["fio2"],
        "tu2": r["tu2"] or "",
        "slot2": r["slot2"],
    } for r in rows]


# ===== импорт файлов =====

def upsert_supply_row(db: Session, point_code: str, supply_date: date, boxes: int):
    db.execute(text("DELETE FROM supplies WHERE point_code=:point_code AND supply_date=:supply_date"), {
        "point_code": point_code, "supply_date": supply_date
    })

    columns = [row[0] for row in db.execute(text("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_name = 'supplies'
        ORDER BY ordinal_position
    """)).fetchall()]

    if 'has_supply' in columns:
        db.execute(text("""
            INSERT INTO supplies (point_code, supply_date, boxes, has_supply)
            VALUES (:point_code, :supply_date, :boxes, :has_supply)
        """), {
            "point_code": point_code,
            "supply_date": supply_date,
            "boxes": boxes,
            "has_supply": True
        })
    else:
        db.execute(text("""
            INSERT INTO supplies (point_code, supply_date, boxes)
            VALUES (:point_code, :supply_date, :boxes)
        """), {
            "point_code": point_code,
            "supply_date": supply_date,
            "boxes": boxes
        })


def import_supplies_xlsx(db: Session, file_obj) -> dict:
    """
    Быстрая загрузка поставок.

    Что изменено:
    - Excel читается один раз;
    - старые поставки по загружаемым точкам/датам удаляются одним SQL-запросом;
    - новые поставки вставляются пачкой через executemany;
    - commit выполняется один раз в конце.

    Формат файла прежний:
    - 1-я колонка: код точки;
    - 1-я строка: даты или номера дней месяца;
    - значения в ячейках: количество коробок;
    - пустые ячейки пропускаются.
    """
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return {"loaded_rows": 0, "loaded_points": 0}

    active = get_active_period()
    report_year = active["year"]
    report_month = active["month"]

    date_columns: list[tuple[int, date]] = []
    for idx, value in enumerate(headers[1:], start=1):
        if value is None:
            continue

        if isinstance(value, datetime):
            date_columns.append((idx, value.date()))
            continue

        if isinstance(value, date):
            date_columns.append((idx, value))
            continue

        raw = str(value).strip()
        match = re.match(r"^(\d{1,2})", raw)
        if not match:
            continue

        day_num = int(match.group(1))
        try:
            supply_date = date(report_year, report_month, day_num)
        except ValueError:
            continue

        date_columns.append((idx, supply_date))

    if not date_columns:
        return {"loaded_rows": 0, "loaded_points": 0}

    supply_rows = []
    loaded_points = set()
    loaded_dates = set()

    for row in rows_iter:
        if not row:
            continue

        point_code = normalize_point_code(row[0] if len(row) > 0 else None)
        if not point_code:
            continue

        row_has_any_supply = False

        for col_idx, supply_date in date_columns:
            raw_boxes = row[col_idx] if col_idx < len(row) else None
            if raw_boxes in (None, ""):
                continue

            try:
                boxes = int(float(raw_boxes))
            except Exception:
                continue

            supply_rows.append({
                "point_code": point_code,
                "supply_date": supply_date,
                "boxes": boxes,
                "has_supply": True,
            })
            loaded_dates.add(supply_date)
            row_has_any_supply = True

        if row_has_any_supply:
            loaded_points.add(point_code)

    if not supply_rows:
        db.commit()
        return {"loaded_rows": 0, "loaded_points": 0}

    columns = [row[0] for row in db.execute(text("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_name = 'supplies'
        ORDER BY ordinal_position
    """)).fetchall()]
    has_supply_column = "has_supply" in columns

    delete_stmt = text("""
        DELETE FROM supplies
        WHERE point_code IN :point_codes
          AND supply_date IN :supply_dates
    """).bindparams(
        bindparam("point_codes", expanding=True),
        bindparam("supply_dates", expanding=True),
    )

    db.execute(delete_stmt, {
        "point_codes": list(loaded_points),
        "supply_dates": list(loaded_dates),
    })

    if has_supply_column:
        insert_stmt = text("""
            INSERT INTO supplies (point_code, supply_date, boxes, has_supply)
            VALUES (:point_code, :supply_date, :boxes, :has_supply)
        """)
        db.execute(insert_stmt, supply_rows)
    else:
        insert_stmt = text("""
            INSERT INTO supplies (point_code, supply_date, boxes)
            VALUES (:point_code, :supply_date, :boxes)
        """)
        db.execute(insert_stmt, [
            {
                "point_code": row["point_code"],
                "supply_date": row["supply_date"],
                "boxes": row["boxes"],
            }
            for row in supply_rows
        ])

    db.commit()
    return {"loaded_rows": len(supply_rows), "loaded_points": len(loaded_points)}


def upsert_rate_row(db: Session, point_code: str, month_key: date, rate_supply: int, rate_no_supply: int, rate_inventory: int, coffee_enabled: bool, coffee_rate: int, pay_lt5: bool):
    db.execute(text("DELETE FROM point_rates WHERE point_code=:point_code AND month_key=:month_key"), {
        "point_code": point_code, "month_key": month_key
    })
    db.execute(text("""
        INSERT INTO point_rates (
            point_code, month_key, rate_supply, rate_no_supply, rate_inventory,
            coffee_enabled, coffee_rate, pay_lt5
        ) VALUES (
            :point_code, :month_key, :rate_supply, :rate_no_supply, :rate_inventory,
            :coffee_enabled, :coffee_rate, :pay_lt5
        )
    """), {
        "point_code": point_code,
        "month_key": month_key,
        "rate_supply": rate_supply,
        "rate_no_supply": rate_no_supply,
        "rate_inventory": rate_inventory,
        "coffee_enabled": coffee_enabled,
        "coffee_rate": coffee_rate,
        "pay_lt5": pay_lt5,
    })


def import_rates_xlsx(db: Session, file_obj, year: int, month: int) -> dict:
    wb = load_workbook(file_obj, data_only=True)
    ws = wb[wb.sheetnames[0]]
    month_key = month_start(year, month)
    loaded_rows = 0

    for row_idx in range(2, ws.max_row + 1):
        point_code = normalize_point_code(ws.cell(row=row_idx, column=1).value)
        if not point_code:
            continue
        rate_supply = int(ws.cell(row=row_idx, column=2).value or 0)
        rate_no_supply = int(ws.cell(row=row_idx, column=3).value or 0)
        rate_inventory = int(ws.cell(row=row_idx, column=4).value or 0)
        coffee_enabled = str(ws.cell(row=row_idx, column=5).value or "").strip().lower() == "да"
        pay_lt5 = str(ws.cell(row=row_idx, column=6).value or "").strip().lower() == "да"
        coffee_rate = int(ws.cell(row=row_idx, column=7).value or 0)
        upsert_rate_row(db, point_code, month_key, rate_supply, rate_no_supply, rate_inventory, coffee_enabled, coffee_rate, pay_lt5)
        loaded_rows += 1

    db.commit()
    return {"loaded_rows": loaded_rows}


def upsert_merchant_row(db: Session, fio: str, last4: str, tu: str):
    fio_clean = str(fio).strip()
    fio_normalized = fio_norm(fio_clean)
    pass_hash = hash_last4(str(last4).strip())
    existing = db.execute(text("SELECT id FROM merchants WHERE fio_norm=:fio_norm LIMIT 1"), {"fio_norm": fio_normalized}).scalar()
    if existing:
        db.execute(text("""
            UPDATE merchants
            SET fio=:fio, fio_norm=:fio_norm, pass_hash=:pass_hash, tu=:tu
            WHERE id=:id
        """), {
            "id": existing,
            "fio": fio_clean,
            "fio_norm": fio_normalized,
            "pass_hash": pass_hash,
            "tu": tu,
        })
    else:
        db.execute(text("""
            INSERT INTO merchants (fio, fio_norm, pass_hash, tu)
            VALUES (:fio, :fio_norm, :pass_hash, :tu)
        """), {
            "fio": fio_clean,
            "fio_norm": fio_normalized,
            "pass_hash": pass_hash,
            "tu": tu,
        })


def import_merchants_xlsx(db: Session, file_obj, tu: str) -> dict:
    wb = load_workbook(file_obj, data_only=True)
    ws = wb[wb.sheetnames[0]]
    loaded_rows = 0
    for row_idx in range(2, ws.max_row + 1):
        fio = ws.cell(row=row_idx, column=1).value
        last4 = ws.cell(row=row_idx, column=2).value
        if not fio or last4 is None:
            continue
        last4_str = re.sub(r"\D", "", str(last4))[-4:]
        if len(last4_str) != 4:
            continue
        upsert_merchant_row(db, str(fio), last4_str, tu)
        loaded_rows += 1
    db.commit()
    return {"loaded_rows": loaded_rows}


def clear_month_data(db: Session, year: int, month: int) -> dict:
    ensure_monthly_submissions_table(db)
    ensure_point_adjustments_table(db)
    start = month_start(year, month)
    end = month_end_exclusive(year, month)
    deleted_visits = db.execute(text("DELETE FROM visits WHERE visit_date >= :start_date AND visit_date < :end_date"), {
        "start_date": start, "end_date": end
    }).rowcount or 0
    deleted_supplies = db.execute(text("DELETE FROM supplies WHERE supply_date >= :start_date AND supply_date < :end_date"), {
        "start_date": start, "end_date": end
    }).rowcount or 0
    deleted_rates = db.execute(text("DELETE FROM point_rates WHERE month_key = :month_key"), {"month_key": start}).rowcount or 0
    deleted_monthly = db.execute(text("DELETE FROM monthly_submissions WHERE month_key = :month_key"), {"month_key": start}).rowcount or 0
    deleted_point_adjustments = db.execute(text("DELETE FROM point_adjustments WHERE month_key = :month_key"), {"month_key": start}).rowcount or 0
    db.commit()
    return {
        "deleted_visits": deleted_visits,
        "deleted_supplies": deleted_supplies,
        "deleted_rates": deleted_rates,
        "deleted_monthly": deleted_monthly,
        "deleted_point_adjustments": deleted_point_adjustments,
    }


def clear_merchants_by_tu(db: Session, tu: str) -> int:
    deleted = db.execute(text("DELETE FROM merchants WHERE tu = :tu"), {"tu": tu}).rowcount or 0
    db.commit()
    return deleted

