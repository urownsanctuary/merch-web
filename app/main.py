
import os
import uuid
import hashlib
from io import BytesIO
from html import escape
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, Depends, HTTPException, Form, UploadFile, File, Cookie
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.db import SessionLocal, engine
from app.services import (
    get_active_period,
    login_user,
    get_merchants_columns,
    normalize_point_code,
    point_has_any_supply_in_month,
    get_supply_boxes_map,
    get_visits_for_month,
    get_merchant_by_fio,
    toggle_day_visit,
    toggle_inventory_visit,
    compute_point_total,
    compute_overall_total,
    days_in_month,
    weekday_of,
    month_title,
    get_monthly_submission,
    upsert_monthly_submission_draft,
    submit_monthly_submission,
    reopen_monthly_submission,
    get_admin_report_rows,
    get_admin_payroll_rows,
    get_intersections_rows,
    get_all_tu_values,
    import_supplies_xlsx,
    import_rates_xlsx,
    import_merchants_xlsx,
    clear_month_data,
    clear_merchants_by_tu,
    get_point_adjustment,
    upsert_point_adjustment,
)

app = FastAPI()

app.mount("/static", StaticFiles(directory="app/static"), name="static")

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

ADMIN_LOGIN = os.getenv("ADMIN_LOGIN", "")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")
SECRET_SALT = os.getenv("SECRET_SALT", "")


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def get_admin_cookie_value() -> str:
    raw = f"{ADMIN_LOGIN}:{ADMIN_PASSWORD}:{SECRET_SALT}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def is_admin_authenticated(admin_auth: Optional[str]) -> bool:
    if not ADMIN_LOGIN or not ADMIN_PASSWORD or not SECRET_SALT:
        return False
    return admin_auth == get_admin_cookie_value()


def style_sheet(ws):
    green_fill = PatternFill("solid", fgColor="E8F5E9")
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)
    ws.freeze_panes = "A2"


def build_excel_response(wb: Workbook, filename: str) -> StreamingResponse:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def base_css():
    return """
    <style>
        @font-face {
            font-family: 'Villula';
            src: url('/static/fonts/villula-regular.ttf') format('truetype');
            font-weight: normal;
            font-style: normal;
        }

        :root {
            --bg: #F6F8F7;
            --card: #FFFFFF;
            --text: #1F2937;
            --muted: #6B7280;
            --line: #D1D5DB;
            --green: #2E7D32;
            --green-dark: #27682A;
            --soft: #EEF4EF;
            --soft-2: #F3F7F3;
            --error: #B91C1C;
            --ok: #166534;
            --shadow: 0 12px 32px rgba(0, 0, 0, 0.08);
        }

        * { box-sizing: border-box; }

        body {
            margin: 0;
            background: var(--bg);
            color: var(--text);
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
            min-height: 100vh;
            padding: 20px;
        }

        .page {
            max-width: 1580px;
            margin: 0 auto;
            min-height: calc(100vh - 40px);
            display: flex;
            align-items: center;
            justify-content: center;
        }

        .card {
            width: 100%;
            max-width: 430px;
            background: var(--card);
            border-radius: 24px;
            padding: 32px 28px;
            box-shadow: var(--shadow);
        }

        .card-wide {
            width: 100%;
            max-width: 1580px;
            background: var(--card);
            border-radius: 24px;
            padding: 22px 20px 28px;
            box-shadow: var(--shadow);
        }

        .brand {
            font-family: 'Villula', -apple-system, sans-serif;
            font-size: 28px;
            line-height: 1;
            color: var(--green);
            margin-bottom: 8px;
        }

        h1 {
            font-family: 'Villula', -apple-system, sans-serif;
            font-size: 34px;
            line-height: 1.05;
            margin: 0 0 8px 0;
            color: var(--text);
        }

        .subtitle {
            color: var(--muted);
            font-size: 15px;
            line-height: 1.45;
            margin-bottom: 18px;
        }

        label {
            display: block;
            margin: 14px 0 6px;
            font-size: 14px;
            font-weight: 700;
            color: var(--text);
        }

        input, textarea, select {
            width: 100%;
            padding: 14px 16px;
            border: 1px solid var(--line);
            border-radius: 14px;
            font-size: 16px;
            background: #fff;
            font-family: inherit;
        }

        textarea {
            resize: vertical;
            min-height: 92px;
        }

        input:focus, textarea:focus, select:focus {
            outline: none;
            border-color: var(--green);
            box-shadow: 0 0 0 3px rgba(46, 125, 50, 0.10);
        }

        .btn {
            display: inline-block;
            width: 100%;
            margin-top: 16px;
            padding: 15px 16px;
            border: none;
            border-radius: 14px;
            background: var(--green);
            color: #fff;
            font-size: 16px;
            font-weight: 800;
            text-align: center;
            text-decoration: none;
            cursor: pointer;
        }

        .btn:hover { background: var(--green-dark); }

        .btn-secondary {
            background: var(--soft);
            color: var(--text);
        }

        .btn-secondary:hover { background: #e3ece3; }

        .btn-danger {
            background: #B91C1C;
            color: #fff;
        }

        .btn-danger:hover {
            background: #991B1B;
        }

        .btn-small {
            margin-top: 12px;
            padding: 12px 14px;
            font-size: 15px;
        }

        .btn-inline {
            width: auto;
            margin-top: 0;
            padding: 12px 16px;
            font-size: 14px;
        }

        .footer {
            margin-top: 18px;
            color: #9CA3AF;
            font-size: 12px;
            text-align: center;
        }

        .back {
            display: inline-block;
            margin-top: 18px;
            color: var(--green);
            text-decoration: none;
            font-weight: 800;
        }

        .hint {
            margin-top: 16px;
            padding: 12px 14px;
            border-radius: 14px;
            background: var(--soft);
            color: var(--muted);
            font-size: 13px;
            line-height: 1.4;
        }

        .error-box {
            margin-top: 16px;
            background: #FEF2F2;
            color: var(--error);
            border-radius: 14px;
            padding: 14px;
            line-height: 1.45;
            font-weight: 700;
        }

        .success-box {
            margin-top: 16px;
            background: #ECFDF3;
            color: var(--ok);
            border-radius: 14px;
            padding: 14px;
            line-height: 1.45;
            font-weight: 700;
        }

        .calendar-head {
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 12px;
            margin-bottom: 14px;
            flex-wrap: wrap;
        }

        .calendar-month {
            font-family: 'Villula', -apple-system, sans-serif;
            font-size: 28px;
            line-height: 1;
        }

        .calendar-meta {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
        }

        .mini-pill {
            background: var(--soft-2);
            border-radius: 999px;
            padding: 8px 12px;
            font-size: 13px;
            color: var(--text);
            font-weight: 700;
        }

        .sum-strip {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 12px;
            margin-bottom: 16px;
        }

        .sum-card {
            background: #F7FBF8;
            border: 1px solid #E5E7EB;
            border-radius: 16px;
            padding: 14px 16px;
        }

        .sum-title {
            color: var(--muted);
            font-size: 13px;
            margin-bottom: 6px;
        }

        .sum-value {
            font-size: 22px;
            font-weight: 900;
        }

        .details-grid {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 12px;
            margin-bottom: 18px;
        }

        .detail-card {
            background: #FAFCFA;
            border: 1px solid #E5E7EB;
            border-radius: 16px;
            padding: 14px 16px;
        }

        .detail-title {
            color: var(--muted);
            font-size: 13px;
            margin-bottom: 8px;
        }

        .detail-line {
            font-size: 15px;
            font-weight: 700;
            line-height: 1.5;
        }

        .calendar-wrap { margin-top: 4px; }

        .weekdays, .calendar-grid {
            display: grid;
            grid-template-columns: repeat(7, 1fr);
            gap: 10px;
        }

        .weekdays { margin-bottom: 10px; }

        .weekday {
            text-align: center;
            font-size: 13px;
            color: var(--muted);
            font-weight: 700;
            padding: 6px 0;
        }

        .day, .day-empty {
            min-height: 90px;
            border-radius: 18px;
            padding: 10px;
        }

        .day {
            background: #F8FAF8;
            border: 1px solid #E5E7EB;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            text-decoration: none;
            color: inherit;
            cursor: pointer;
        }

        .day:hover {
            border-color: var(--green);
            box-shadow: 0 0 0 2px rgba(46, 125, 50, 0.06);
        }

        .day-empty { background: transparent; }

        .day-disabled {
            opacity: 0.65;
            cursor: default;
            pointer-events: none;
        }

        .day-number {
            font-size: 18px;
            font-weight: 800;
        }

        .day-badges {
            display: flex;
            flex-wrap: wrap;
            gap: 6px;
            margin-top: 10px;
        }

        .badge {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            padding: 4px 8px;
            border-radius: 999px;
            font-size: 11px;
            font-weight: 800;
            line-height: 1;
            min-width: 22px;
            height: 22px;
        }

        .badge-supply {
            background: #2E7D32;
            color: #fff;
            border-radius: 6px;
        }

        .badge-day {
            background: #DBEAFE;
            color: #1D4ED8;
        }

        .badge-inv {
            background: #FCE7F3;
            color: #BE185D;
        }

        .legend {
            margin-top: 18px;
            display: flex;
            flex-wrap: wrap;
            gap: 10px;
        }

        .legend-item {
            background: var(--soft);
            border-radius: 999px;
            padding: 8px 12px;
            font-size: 13px;
            color: var(--text);
            font-weight: 700;
        }

        .calendar-note {
            margin-top: 14px;
            color: var(--muted);
            line-height: 1.5;
            font-size: 14px;
        }

        details.point-detail {
            margin-top: 10px;
            border: 1px solid #E5E7EB;
            border-radius: 14px;
            background: #FAFCFA;
            padding: 10px 14px;
        }

        details.point-detail summary {
            cursor: pointer;
            font-weight: 800;
            list-style: none;
        }

        details.point-detail summary::-webkit-details-marker {
            display: none;
        }

        .summary-content {
            margin-top: 10px;
            color: var(--text);
            line-height: 1.6;
        }

        .filter-grid {
            display: grid;
            grid-template-columns: 140px 140px 240px 220px;
            gap: 12px;
            margin-bottom: 16px;
            align-items: end;
        }

        .table-wrap {
            border: 1px solid #E5E7EB;
            border-radius: 16px;
            background: #fff;
            overflow: visible;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
        }

        th, td {
            padding: 10px 8px;
            border-bottom: 1px solid #E5E7EB;
            text-align: left;
            vertical-align: top;
            font-size: 12px;
            word-break: break-word;
        }

        th {
            background: #F7FBF8;
            font-weight: 800;
        }

        .admin-actions {
            display: flex;
            gap: 12px;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 16px;
            flex-wrap: wrap;
        }

        .admin-export-buttons {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
            margin-top: 14px;
        }

        .data-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 16px;
        }

        @media (max-width: 960px) {
            .page { align-items: flex-start; }
            .card-wide { padding: 18px 14px 24px; }
            .sum-strip, .details-grid, .filter-grid, .data-grid { grid-template-columns: 1fr; }
            .weekdays, .calendar-grid { gap: 8px; }
            .day, .day-empty {
                min-height: 80px;
                border-radius: 14px;
                padding: 8px;
            }
            .day-number { font-size: 16px; }
            h1 { font-size: 30px; }
            .brand { font-size: 24px; }
            .calendar-month { font-size: 24px; }
            .table-wrap { overflow-x: auto; }
            table { min-width: 1200px; }
        }
    </style>
    """


@app.get("/")
def root():
    return RedirectResponse(url="/login-page")


@app.get("/db-check")
def db_check():
    with engine.connect() as conn:
        conn.execute(text("SELECT 1"))
    return {"status": "ok", "db": "connected"}


@app.get("/active-period")
def active_period():
    return get_active_period()


@app.get("/debug/merchants-columns")
def merchants_columns(db: Session = Depends(get_db)):
    cols = get_merchants_columns(db)
    return {"table": "merchants", "columns": cols}


@app.post("/login")
def login_api(fio: str, last4: str, db: Session = Depends(get_db)):
    user = login_user(db, fio, last4)

    if not user:
        raise HTTPException(status_code=401, detail="ÐÐµÐ²ÐµÑÐ½ÑÐµ Ð´Ð°Ð½Ð½ÑÐµ")

    return {"status": "ok", "active_period": get_active_period(), "user": user}


@app.get("/login-page", response_class=HTMLResponse)
def login_page():
    period = get_active_period()
    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>ÐÐºÑÑÐÐ¸Ð»Ð»</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <div class="brand">ÐÐºÑÑÐÐ¸Ð»Ð»</div>
            <h1>Ð¡Ð²ÐµÑÐºÐ¸ Ð¼ÐµÑÑÐµÐ½Ð´Ð°Ð¹Ð·ÐµÑÐ¾Ð²</h1>
            <div class="subtitle">ÐÐ²ÐµÐ´Ð¸ÑÐµ Ð¤ÐÐ Ð¸ Ð¿Ð¾ÑÐ»ÐµÐ´Ð½Ð¸Ðµ 4 ÑÐ¸ÑÑÑ ÑÐµÐ»ÐµÑÐ¾Ð½Ð°</div>

            <form method="post" action="/login-page">
                <label for="fio">Ð¤ÐÐ</label>
                <input id="fio" name="fio" type="text" placeholder="ÐÐ²Ð°Ð½Ð¾Ð² ÐÐ²Ð°Ð½ ÐÐ²Ð°Ð½Ð¾Ð²Ð¸Ñ" required />

                <label for="last4">ÐÐ¾ÑÐ»ÐµÐ´Ð½Ð¸Ðµ 4 ÑÐ¸ÑÑÑ ÑÐµÐ»ÐµÑÐ¾Ð½Ð°</label>
                <input id="last4" name="last4" type="text" inputmode="numeric" maxlength="4" placeholder="1234" required />

                <button class="btn" type="submit">ÐÐ¾Ð¹ÑÐ¸</button>
            </form>

            <div class="hint">Ð¡ÐµÐ¹ÑÐ°Ñ Ð¾ÑÐºÑÑÑ Ð¿ÐµÑÐ¸Ð¾Ð´ Ð·Ð° {month_title(period["year"], period["month"])}.</div>

            <div class="footer">ÐÐµÐ±-Ð²ÐµÑÑÐ¸Ñ ÑÐ²ÐµÑÐ¾Ðº Ð¼ÐµÑÑÐµÐ½Ð´Ð°Ð¹Ð·ÐµÑÐ¾Ð²</div>
        </div>
    </div>
</body>
</html>
"""


@app.post("/login-page", response_class=HTMLResponse)
def login_submit(
    fio: str = Form(...),
    last4: str = Form(...),
    db: Session = Depends(get_db)
):
    user = login_user(db, fio, last4)

    if not user:
        return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>ÐÑÐ¸Ð±ÐºÐ° Ð²ÑÐ¾Ð´Ð°</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <h1>ÐÑÐ¸Ð±ÐºÐ° Ð²ÑÐ¾Ð´Ð°</h1>
            <div class="error-box">ÐÐµÐ²ÐµÑÐ½ÑÐµ Ð´Ð°Ð½Ð½ÑÐµ. ÐÑÐ¾Ð²ÐµÑÑÑÐµ Ð¤ÐÐ Ð¸ Ð¿Ð¾ÑÐ»ÐµÐ´Ð½Ð¸Ðµ 4 ÑÐ¸ÑÑÑ ÑÐµÐ»ÐµÑÐ¾Ð½Ð°.</div>
            <a class="back" href="/login-page">â ÐÐ¾Ð¿ÑÐ¾Ð±Ð¾Ð²Ð°ÑÑ ÑÐ½Ð¾Ð²Ð°</a>
        </div>
    </div>
</body>
</html>
"""

    return RedirectResponse(url=f"/menu-page?fio={user['fio']}", status_code=303)


@app.get("/menu-page", response_class=HTMLResponse)
def menu_page(fio: str = "", db: Session = Depends(get_db)):
    period = get_active_period()
    merchant = get_merchant_by_fio(db, fio)
    overall = {"total": 0}
    if merchant:
        overall = compute_overall_total(db, merchant["id"], period["year"], period["month"])

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>ÐÐ»Ð°Ð²Ð½Ð¾Ðµ Ð¼ÐµÐ½Ñ</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <div class="brand">ÐÐºÑÑÐÐ¸Ð»Ð»</div>
            <h1>ÐÐ»Ð°Ð²Ð½Ð¾Ðµ Ð¼ÐµÐ½Ñ</h1>
            <div class="subtitle">{escape(fio)}</div>
            <div class="hint">Ð¡ÐµÐ¹ÑÐ°Ñ Ð¾ÑÐºÑÑÑ Ð¿ÐµÑÐ¸Ð¾Ð´ Ð·Ð° {month_title(period["year"], period["month"])}.</div>

            <div class="sum-card" style="margin-top: 18px;">
                <div class="sum-title">ÐÐ±ÑÐ°Ñ ÑÑÐ¼Ð¼Ð° Ð·Ð° Ð¼ÐµÑÑÑ</div>
                <div class="sum-value">{overall["total"]} â½</div>
            </div>

            <a class="btn" href="/point-page?fio={escape(fio)}">ÐÐ°Ð¿Ð¾Ð»Ð½Ð¸ÑÑ ÑÐ²ÐµÑÐºÑ</a>
            <a class="btn btn-secondary" href="/summary-page?fio={escape(fio)}">ÐÐ¾Ñ ÑÑÐ¼Ð¼Ð°</a>
            <a class="btn btn-secondary" href="/monthly-submit-page?fio={escape(fio)}">ÐÑÐ¿ÑÐ°Ð²Ð¸ÑÑ ÑÐ²ÐµÑÐºÑ Ð·Ð° Ð¼ÐµÑÑÑ</a>
        </div>
    </div>
</body>
</html>
"""


@app.get("/point-page", response_class=HTMLResponse)
def point_page(fio: str = ""):
    period = get_active_period()

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>ÐÑÐ±Ð¾Ñ ÑÐ¾ÑÐºÐ¸</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <div class="brand">ÐÐºÑÑÐÐ¸Ð»Ð»</div>
            <h1>ÐÑÐ±Ð¾Ñ ÑÐ¾ÑÐºÐ¸</h1>
            <div class="subtitle">{escape(fio)}</div>

            <div class="hint" style="margin-top: 0; margin-bottom: 18px;">
                Ð¡Ð²ÐµÑÐºÐ° Ð·Ð°Ð¿Ð¾Ð»Ð½ÑÐµÑÑÑ Ð·Ð° {month_title(period["year"], period["month"])}.
            </div>

            <form method="post" action="/point-page">
                <input type="hidden" name="fio" value="{escape(fio)}" />

                <label for="point_code">ÐÐ¾Ð¼ÐµÑ ÑÐ¾ÑÐºÐ¸</label>
                <input id="point_code" name="point_code" type="text" placeholder="2674" required />

                <button class="btn" type="submit">ÐÑÐ¾Ð´Ð¾Ð»Ð¶Ð¸ÑÑ</button>
            </form>

            <a class="back" href="/menu-page?fio={escape(fio)}">â ÐÐ°Ð·Ð°Ð´</a>
        </div>
    </div>
</body>
</html>
"""


@app.post("/point-page", response_class=HTMLResponse)
def point_submit(
    fio: str = Form(...),
    point_code: str = Form(...),
    db: Session = Depends(get_db)
):
    period = get_active_period()
    point_code = normalize_point_code(point_code)

    if not point_code or len(point_code) < 3:
        return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>ÐÑÐ¸Ð±ÐºÐ°</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <h1>ÐÑÐ¸Ð±ÐºÐ°</h1>
            <div class="error-box">ÐÐ¾Ð¼ÐµÑ ÑÐ¾ÑÐºÐ¸ ÑÐ»Ð¸ÑÐºÐ¾Ð¼ ÐºÐ¾ÑÐ¾ÑÐºÐ¸Ð¹.</div>
            <a class="back" href="/point-page?fio={escape(fio)}">â ÐÐ°Ð·Ð°Ð´</a>
        </div>
    </div>
</body>
</html>
"""

    has_supply = point_has_any_supply_in_month(db, point_code, period["year"], period["month"])

    if not has_supply:
        return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Ð¢Ð¾ÑÐºÐ° Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½Ð°</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <h1>Ð¢Ð¾ÑÐºÐ° Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½Ð°</h1>
            <div class="error-box">
                Ð Ð¿ÐµÑÐ¸Ð¾Ð´Ðµ {month_title(period["year"], period["month"])} Ð¿Ð¾ ÑÐ¾ÑÐºÐµ {escape(point_code)} Ð½ÐµÑ Ð¿Ð¾ÑÑÐ°Ð²Ð¾Ðº.
                <br><br>
                ÐÑÐ¾Ð²ÐµÑÑÑÐµ Ð½Ð¾Ð¼ÐµÑ ÑÐ¾ÑÐºÐ¸ Ð¸Ð»Ð¸ Ð¾Ð±ÑÐ°ÑÐ¸ÑÐµÑÑ Ðº ÑÐ¿ÑÐ°Ð²Ð»ÑÑÑÐµÐ¼Ñ.
            </div>
            <a class="back" href="/point-page?fio={escape(fio)}">â ÐÐ¾Ð¿ÑÐ¾Ð±Ð¾Ð²Ð°ÑÑ ÑÐ½Ð¾Ð²Ð°</a>
        </div>
    </div>
</body>
</html>
"""

    return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)


def build_day_href(fio: str, point_code: str, y: int, m: int, day: int, is_submitted: bool) -> str:
    if is_submitted:
        return "#"
    wd = weekday_of(y, m, day)
    if wd in (4, 5):
        return f"/day-action-page?fio={escape(fio)}&point_code={escape(point_code)}&day={day}"
    return f"/toggle-day?fio={escape(fio)}&point_code={escape(point_code)}&day={day}"


def build_calendar_html(
    fio: str,
    point_code: str,
    y: int,
    m: int,
    boxes_map: dict[int, int],
    visits: dict[int, set[str]],
    is_submitted: bool
) -> str:
    dim = days_in_month(y, m)
    first_wd = weekday_of(y, m, 1)
    weekdays = ["ÐÐ½", "ÐÑ", "Ð¡Ñ", "Ð§Ñ", "ÐÑ", "Ð¡Ð±", "ÐÑ"]

    html = '<div class="weekdays">'
    for wd in weekdays:
        html += f'<div class="weekday">{wd}</div>'
    html += '</div>'

    html += '<div class="calendar-grid">'

    for _ in range(first_wd):
        html += '<div class="day-empty"></div>'

    for day in range(1, dim + 1):
        boxes = boxes_map.get(day, 0)
        day_visits = visits.get(day, set())

        badges = ""
        if boxes > 0:
            badges += '<span class="badge badge-supply">Ð</span>'
        if "DAY" in day_visits:
            badges += '<span class="badge badge-day">Ð</span>'
        if "FULL_INVENT" in day_visits:
            badges += '<span class="badge badge-inv">Ð</span>'

        href = build_day_href(fio, point_code, y, m, day, is_submitted)
        cls = "day day-disabled" if is_submitted else "day"

        html += f"""
        <a class="{cls}" href="{href}">
            <div class="day-number">{day}</div>
            <div class="day-badges">{badges}</div>
        </a>
        """

    html += '</div>'
    return html


@app.get("/calendar-page", response_class=HTMLResponse)
def calendar_page(
    fio: str,
    point_code: str,
    saved: str = "",
    db: Session = Depends(get_db)
):
    period = get_active_period()
    y = period["year"]
    m = period["month"]

    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    point_code = normalize_point_code(point_code)

    overall = compute_overall_total(db, merchant["id"], y, m)
    monthly_submitted = overall["submission_status"] == "submitted"

    boxes_map = get_supply_boxes_map(db, point_code, y, m)
    visits = get_visits_for_month(db, merchant["id"], point_code, y, m)
    point_total = compute_point_total(db, merchant["id"], point_code, y, m)
    point_adj = get_point_adjustment(db, merchant["id"], point_code, y, m) or {}

    calendar_html = build_calendar_html(
        fio=fio,
        point_code=point_code,
        y=y,
        m=m,
        boxes_map=boxes_map,
        visits=visits,
        is_submitted=monthly_submitted
    )

    info_box = ""
    if saved == "1":
        info_box = "<div class=\"success-box\">ÐÐ°Ð½Ð½ÑÐµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ ÑÐ¾ÑÑÐ°Ð½ÐµÐ½Ñ.</div>"

    point_receipt_link = ""
    if point_total["reimb_receipt"]:
        point_receipt_link = f"<div class='hint' style='margin-top:10px'>Ð§ÐµÐº Ð¿Ð¾ ÑÐ¾ÑÐºÐµ: <a href='/{point_total['reimb_receipt']}' target='_blank'>Ð¾ÑÐºÑÑÑÑ ÑÐ°Ð¹Ð»</a></div>"

    point_form = ""
    if not monthly_submitted:
        point_form = f"""
            <form method="post" action="/save-point-adjustment" enctype="multipart/form-data" class="detail-card" style="margin-top:18px;">
                <input type="hidden" name="fio" value="{escape(fio)}" />
                <input type="hidden" name="point_code" value="{escape(point_code)}" />

                <div class="detail-title">ÐÑÐ¸Ð¼ÐµÑÐ°Ð½Ð¸Ðµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ</div>
                <label for="note_amount">Ð¡ÑÐ¼Ð¼Ð°, â½</label>
                <input id="note_amount" name="note_amount" type="number" min="0" value="{point_total['note_amount']}" placeholder="ÐÐ°Ð¿ÑÐ¸Ð¼ÐµÑ: 1500" />

                <label for="note_comment">ÐÐ¾Ð¼Ð¼ÐµÐ½ÑÐ°ÑÐ¸Ð¹</label>
                <input id="note_comment" name="note_comment" type="text" value="{escape(point_total['note_comment'])}" placeholder="ÐÐ°Ð¿ÑÐ¸Ð¼ÐµÑ: ÐÐ°ÐºÑÑÑÐ¸Ðµ ÑÐ¾ÑÐºÐ¸" />

                <div class="detail-title" style="margin-top:18px;">ÐÐ¾Ð·Ð¼ÐµÑÐµÐ½Ð¸Ðµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ</div>
                <label for="reimb_amount">Ð¡ÑÐ¼Ð¼Ð°, â½</label>
                <input id="reimb_amount" name="reimb_amount" type="number" min="0" value="{point_total['reimb_amount']}" placeholder="ÐÐ°Ð¿ÑÐ¸Ð¼ÐµÑ: 150" />

                <label for="reimb_comment">ÐÐ¾Ð¼Ð¼ÐµÐ½ÑÐ°ÑÐ¸Ð¹</label>
                <input id="reimb_comment" name="reimb_comment" type="text" value="{escape(point_total['reimb_comment'])}" placeholder="ÐÐ°Ð¿ÑÐ¸Ð¼ÐµÑ: ÐÐ¾ÐºÑÐ¿ÐºÐ° Ð¿Ð°ÐºÐµÑÐ¾Ð²" />

                <label for="reimb_receipt">Ð§ÐµÐº</label>
                <input id="reimb_receipt" name="reimb_receipt" type="file" accept=".jpg,.jpeg,.png,.pdf,.webp" />
                <div class="hint" style="margin-top:10px;">ÐÑÐ»Ð¸ ÑÐºÐ°Ð·Ð°Ð½Ð¾ Ð²Ð¾Ð·Ð¼ÐµÑÐµÐ½Ð¸Ðµ, ÑÐµÐº Ð¾Ð±ÑÐ·Ð°ÑÐµÐ»ÐµÐ½.</div>
                {point_receipt_link}

                <button class="btn btn-secondary" type="submit">Ð¡Ð¾ÑÑÐ°Ð½Ð¸ÑÑ Ð´Ð°Ð½Ð½ÑÐµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ</button>
            </form>
        """

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>ÐÐ°Ð»ÐµÐ½Ð´Ð°ÑÑ</title>
    {base_css()}
    <script>
    document.addEventListener('DOMContentLoaded', function() {{
        const savedY = sessionStorage.getItem('calendarScrollY');
        if (savedY) {{
            window.scrollTo(0, parseInt(savedY, 10));
            sessionStorage.removeItem('calendarScrollY');
        }}
        document.querySelectorAll('.day').forEach(el => {{
            el.addEventListener('click', function() {{
                sessionStorage.setItem('calendarScrollY', String(window.scrollY));
            }});
        }});
    }});
    </script>
</head>
<body>
    <div class="page">
        <div class="card-wide">
            <div class="calendar-head">
                <div>
                    <div class="brand">ÐÐºÑÑÐÐ¸Ð»Ð»</div>
                    <div class="calendar-month">{month_title(y, m)}</div>
                </div>

                <div class="calendar-meta">
                    <div class="mini-pill">Ð¢Ð¾ÑÐºÐ°: {escape(point_code)}</div>
                    <div class="mini-pill">{escape(fio)}</div>
                    <div class="mini-pill">ÐÐ: {"ÐÐ°" if point_total["coffee_enabled"] else "ÐÐµÑ"}</div>
                    <div class="mini-pill">ÐÐµÑÑÑÐ½Ð°Ñ ÑÐ²ÐµÑÐºÐ°: {"ÐÑÐ¿ÑÐ°Ð²Ð»ÐµÐ½Ð°" if monthly_submitted else "Ð§ÐµÑÐ½Ð¾Ð²Ð¸Ðº"}</div>
                </div>
            </div>

            {info_box}

            <div class="sum-strip">
                <div class="sum-card">
                    <div class="sum-title">Ð¡ÑÐ¼Ð¼Ð° Ð¿Ð¾ ÑÐ¾ÑÐºÐµ</div>
                    <div class="sum-value">{point_total["total"]} â½</div>
                </div>

                <div class="sum-card">
                    <div class="sum-title">ÐÐ±ÑÐ°Ñ ÑÑÐ¼Ð¼Ð° Ð·Ð° Ð¼ÐµÑÑÑ</div>
                    <div class="sum-value">{overall["total"]} â½</div>
                </div>
            </div>

            <div class="details-grid">
                <div class="detail-card">
                    <div class="detail-title">ÐÑÑÐ¾Ð´Ñ Ñ Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¾Ð¹</div>
                    <div class="detail-line">{point_total["cnt_supply"]} Ã {point_total["rate_supply"]} â½ = {point_total["sum_supply"]} â½</div>
                </div>

                <div class="detail-card">
                    <div class="detail-title">ÐÑÑÐ¾Ð´Ñ Ð±ÐµÐ· Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¸</div>
                    <div class="detail-line">{point_total["cnt_no_supply"]} Ã {point_total["rate_no_supply"]} â½ = {point_total["sum_no_supply"]} â½</div>
                </div>

                <div class="detail-card">
                    <div class="detail-title">ÐÐ¾Ð»Ð½ÑÐµ Ð¸Ð½Ð²ÐµÐ½ÑÑ</div>
                    <div class="detail-line">{point_total["cnt_full_inv"]} Ã {point_total["rate_inventory"]} â½ = {point_total["sum_inventory"]} â½</div>
                </div>

                <div class="detail-card">
                    <div class="detail-title">ÐÐ¾ÑÐµÐ¼Ð°ÑÐ¸Ð½Ð°</div>
                    <div class="detail-line">{point_total["coffee_cnt"]} Ã {point_total["coffee_rate"]} â½ = {point_total["coffee_sum"]} â½</div>
                </div>
            </div>

            <div class="details-grid">
                <div class="detail-card">
                    <div class="detail-title">ÐÑÐ¸Ð¼ÐµÑÐ°Ð½Ð¸Ðµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ</div>
                    <div class="detail-line">{point_total["note_amount"]} â½</div>
                    <div class="calendar-note">{escape(point_total["note_comment"]) if point_total["note_comment"] else "â"}</div>
                </div>

                <div class="detail-card">
                    <div class="detail-title">ÐÐ¾Ð·Ð¼ÐµÑÐµÐ½Ð¸Ðµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ</div>
                    <div class="detail-line">{point_total["reimb_amount"]} â½</div>
                    <div class="calendar-note">{escape(point_total["reimb_comment"]) if point_total["reimb_comment"] else "â"}</div>
                </div>
            </div>

            <div class="calendar-wrap">
                {calendar_html}
            </div>

            <div class="legend">
                <div class="legend-item">Ð â Ð±ÑÐ»Ð° Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ°</div>
                <div class="legend-item">Ð â Ð¾ÑÐ¼ÐµÑÐµÐ½ Ð²ÑÑÐ¾Ð´</div>
                <div class="legend-item">Ð â Ð¿Ð¾Ð»Ð½ÑÐ¹ Ð¸Ð½Ð²ÐµÐ½Ñ</div>
            </div>

            <div class="calendar-note">
                Ð Ð¾Ð±ÑÑÐ½ÑÐµ Ð´Ð½Ð¸ Ð½Ð°Ð¶Ð°ÑÐ¸Ðµ Ð¿Ð¾ Ð´Ð½Ñ ÑÑÐ°Ð·Ñ ÑÑÐ°Ð²Ð¸Ñ Ð¸Ð»Ð¸ ÑÐ±Ð¸ÑÐ°ÐµÑ Ð²ÑÑÐ¾Ð´.
                Ð Ð¿ÑÑÐ½Ð¸ÑÑ Ð¸ ÑÑÐ±Ð±Ð¾ÑÑ Ð¾ÑÐºÑÑÐ²Ð°ÐµÑÑÑ Ð²ÑÐ±Ð¾Ñ: Ð²ÑÑÐ¾Ð´ Ð¸Ð»Ð¸ Ð¿Ð¾Ð»Ð½ÑÐ¹ Ð¸Ð½Ð²ÐµÐ½Ñ.
            </div>

            <div class="calendar-note">
                ÐÐ¾ÑÑÐ°Ð²ÐºÐ¸ Ð´Ð¾ 5 ÐºÐ¾ÑÐ¾Ð±Ð¾Ðº Ð½Ðµ Ð¾Ð¿Ð»Ð°ÑÐ¸Ð²Ð°ÑÑÑÑ.
            </div>

            {point_form}

            <div class="admin-export-buttons" style="margin-top:18px;">
                <a class="btn btn-secondary btn-inline" href="/point-page?fio={escape(fio)}">Ð¡Ð»ÐµÐ´ÑÑÑÐ°Ñ ÑÐ¾ÑÐºÐ°</a>
                <a class="btn btn-secondary btn-inline" href="/summary-page?fio={escape(fio)}">ÐÐ¾Ñ ÑÑÐ¼Ð¼Ð°</a>
                <a class="btn btn-secondary btn-inline" href="/monthly-submit-page?fio={escape(fio)}">ÐÑÐ¿ÑÐ°Ð²Ð¸ÑÑ ÑÐ²ÐµÑÐºÑ Ð·Ð° Ð¼ÐµÑÑÑ</a>
            </div>

            <a class="back" href="/menu-page?fio={escape(fio)}">â ÐÐ° Ð³Ð»Ð°Ð²Ð½ÑÐ¹ ÑÐºÑÐ°Ð½</a>
        </div>
    </div>
</body>
</html>
"""


@app.post("/save-point-adjustment")
async def save_point_adjustment(
    fio: str = Form(...),
    point_code: str = Form(...),
    note_amount: int = Form(0),
    note_comment: str = Form(""),
    reimb_amount: int = Form(0),
    reimb_comment: str = Form(""),
    reimb_receipt: UploadFile | None = File(None),
    db: Session = Depends(get_db)
):
    period = get_active_period()
    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    receipt_path = None
    if reimb_receipt and reimb_receipt.filename:
        ext = Path(reimb_receipt.filename).suffix.lower()
        filename = f"{uuid.uuid4().hex}{ext}"
        filepath = UPLOAD_DIR / filename
        content = await reimb_receipt.read()
        filepath.write_bytes(content)
        receipt_path = f"uploads/{filename}"

    upsert_point_adjustment(
        db=db,
        merchant_id=merchant["id"],
        point_code=normalize_point_code(point_code),
        y=period["year"],
        m=period["month"],
        note_amount=max(0, int(note_amount or 0)),
        note_comment=note_comment or "",
        reimb_amount=max(0, int(reimb_amount or 0)),
        reimb_comment=reimb_comment or "",
        reimb_receipt=receipt_path,
    )

    return RedirectResponse(
        url=f"/calendar-page?fio={escape(fio)}&point_code={escape(normalize_point_code(point_code))}&saved=1",
        status_code=303
    )



@app.get("/monthly-submit-page", response_class=HTMLResponse)
def monthly_submit_page(
    fio: str,
    submitted: str = "",
    reopened: str = "",
    db: Session = Depends(get_db)
):
    period = get_active_period()
    y = period["year"]
    m = period["month"]

    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    overall = compute_overall_total(db, merchant["id"], y, m)
    monthly_submitted = overall["submission_status"] == "submitted"

    info_box = ""
    if submitted == "1":
        info_box += "<div class='success-box'>ÐÐµÑÑÑÐ½Ð°Ñ ÑÐ²ÐµÑÐºÐ° Ð¾ÑÐ¿ÑÐ°Ð²Ð»ÐµÐ½Ð°.</div>"
    if reopened == "1":
        info_box += "<div class='success-box'>ÐÐµÑÑÑÐ½Ð°Ñ ÑÐ²ÐµÑÐºÐ° ÑÐ°Ð·Ð±Ð»Ð¾ÐºÐ¸ÑÐ¾Ð²Ð°Ð½Ð° Ð´Ð»Ñ ÑÐµÐ´Ð°ÐºÑÐ¸ÑÐ¾Ð²Ð°Ð½Ð¸Ñ.</div>"

    points_html = ""
    if overall["per_point_details"]:
        for point_code, d in overall["per_point_details"].items():
            points_html += f"""
            <details class="point-detail">
                <summary>{escape(point_code)} â {d["total"]} â½</summary>
                <div class="summary-content">
                    <div>Ð¡ Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¾Ð¹: {d["cnt_supply"]} Ã {d["rate_supply"]} â½ = {d["sum_supply"]} â½</div>
                    <div>ÐÐµÐ· Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¸: {d["cnt_no_supply"]} Ã {d["rate_no_supply"]} â½ = {d["sum_no_supply"]} â½</div>
                    <div>ÐÐ¾Ð»Ð½ÑÐ¹ Ð¸Ð½Ð²ÐµÐ½Ñ: {d["cnt_full_inv"]} Ã {d["rate_inventory"]} â½ = {d["sum_inventory"]} â½</div>
                    <div>ÐÐ¾ÑÐµÐ¼Ð°ÑÐ¸Ð½Ð°: {d["coffee_cnt"]} Ã {d["coffee_rate"]} â½ = {d["coffee_sum"]} â½</div>
                    <div>ÐÑÐ¸Ð¼ÐµÑÐ°Ð½Ð¸Ðµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ: {d["note_amount"]} â½ â {escape(d["note_comment"]) if d["note_comment"] else "â"}</div>
                    <div>ÐÐ¾Ð·Ð¼ÐµÑÐµÐ½Ð¸Ðµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ: {d["reimb_amount"]} â½ â {escape(d["reimb_comment"]) if d["reimb_comment"] else "â"}</div>
                    <div>Ð§ÐµÐº Ð¿Ð¾ Ð²Ð¾Ð·Ð¼ÐµÑÐµÐ½Ð¸Ñ: {f"<a href='/{d['reimb_receipt']}' target='_blank'>Ð¾ÑÐºÑÑÑÑ</a>" if d["reimb_receipt"] else "â"}</div>
                </div>
            </details>
            """
    else:
        points_html = "<div class='hint'>Ð ÑÑÐ¾Ð¼ Ð¼ÐµÑÑÑÐµ Ð¿Ð¾ÐºÐ° Ð½ÐµÑ Ð¾ÑÐ¼ÐµÑÐµÐ½Ð½ÑÑ ÑÐ¾ÑÐµÐº.</div>"

    action_block = ""
    if monthly_submitted:
        action_block = f"""
        <div class="detail-card" style="margin-top:18px;">
            <div class="detail-title">Ð¡ÑÐ°ÑÑÑ</div>
            <div class="detail-line">Ð¡Ð²ÐµÑÐºÐ° Ð·Ð° Ð¼ÐµÑÑÑ Ð¾ÑÐ¿ÑÐ°Ð²Ð»ÐµÐ½Ð°</div>
            <a class="btn btn-secondary" href="/reopen-monthly-submission?fio={escape(fio)}">Ð ÐµÐ´Ð°ÐºÑÐ¸ÑÐ¾Ð²Ð°ÑÑ ÑÐ²ÐµÑÐºÑ</a>
        </div>
        """
    else:
        action_block = f"""
        <form method="post" action="/submit-monthly-submission">
            <input type="hidden" name="fio" value="{escape(fio)}" />
            <button class="btn" type="submit">ÐÑÐ¿ÑÐ°Ð²Ð¸ÑÑ ÑÐ²ÐµÑÐºÑ Ð·Ð° Ð¼ÐµÑÑÑ</button>
        </form>
        """

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>ÐÑÐ¿ÑÐ°Ð²ÐºÐ° Ð¼ÐµÑÑÑÐ½Ð¾Ð¹ ÑÐ²ÐµÑÐºÐ¸</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card-wide">
            <div class="brand">ÐÐºÑÑÐÐ¸Ð»Ð»</div>
            <h1>ÐÑÐ¿ÑÐ°Ð²Ð¸ÑÑ ÑÐ²ÐµÑÐºÑ Ð·Ð° Ð¼ÐµÑÑÑ</h1>
            <div class="subtitle">{escape(fio)} Â· {month_title(y, m)}</div>

            {info_box}

            <div class="sum-strip">
                <div class="sum-card">
                    <div class="sum-title">Ð¡ÑÐ¼Ð¼Ð° Ð¿Ð¾ ÑÐ¾ÑÐºÐ°Ð¼</div>
                    <div class="sum-value">{sum(overall["per_point"].values())} â½</div>
                </div>

                <div class="sum-card">
                    <div class="sum-title">ÐÑÐ¾Ð³ Ð·Ð° Ð¼ÐµÑÑÑ</div>
                    <div class="sum-value">{overall["total"]} â½</div>
                </div>
            </div>

            <div class="hint">
                ÐÐ° ÑÑÐ¾Ð¹ ÑÑÑÐ°Ð½Ð¸ÑÐµ Ð±Ð¾Ð»ÑÑÐµ Ð½ÐµÑ Ð¿Ð¾Ð»ÐµÐ¹ Ð¿ÑÐ¸Ð¼ÐµÑÐ°Ð½Ð¸Ñ Ð¸ Ð²Ð¾Ð·Ð¼ÐµÑÐµÐ½Ð¸Ñ Ð·Ð° Ð¼ÐµÑÑÑ.
                ÐÐ½Ð¸ Ð·Ð°Ð¿Ð¾Ð»Ð½ÑÑÑÑÑ Ð¾ÑÐ´ÐµÐ»ÑÐ½Ð¾ Ð²Ð½ÑÑÑÐ¸ ÐºÐ°Ð¶Ð´Ð¾Ð¹ ÑÐ¾ÑÐºÐ¸.
            </div>

            {points_html}

            {action_block}

            <div class="admin-export-buttons" style="margin-top:18px;">
                <a class="btn btn-secondary btn-inline" href="/point-page?fio={escape(fio)}">ÐÐµÑÐµÐ¹ÑÐ¸ Ðº Ð´ÑÑÐ³Ð¾Ð¹ ÑÐ¾ÑÐºÐµ</a>
                <a class="btn btn-secondary btn-inline" href="/summary-page?fio={escape(fio)}">ÐÐ¾Ñ ÑÑÐ¼Ð¼Ð°</a>
                <a class="btn btn-secondary btn-inline" href="/menu-page?fio={escape(fio)}">ÐÐ»Ð°Ð²Ð½ÑÐ¹ ÑÐºÑÐ°Ð½</a>
            </div>
        </div>
    </div>
</body>
</html>
"""


@app.post("/submit-monthly-submission")
async def submit_monthly_submission_route(
    fio: str = Form(...),
    db: Session = Depends(get_db)
):
    period = get_active_period()
    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    submit_monthly_submission(db, merchant["id"], period["year"], period["month"])

    return RedirectResponse(
        url=f"/monthly-submit-page?fio={escape(fio)}&submitted=1",
        status_code=303
    )


@app.get("/reopen-monthly-submission")
def reopen_monthly_submission_route(
    fio: str,
    db: Session = Depends(get_db)
):
    period = get_active_period()
    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    reopen_monthly_submission(db, merchant["id"], period["year"], period["month"])

    return RedirectResponse(
        url=f"/monthly-submit-page?fio={escape(fio)}&reopened=1",
        status_code=303
    )

@app.get("/day-action-page", response_class=HTMLResponse)
def day_action_page(
    fio: str,
    point_code: str,
    day: int,
    db: Session = Depends(get_db)
):
    period = get_active_period()
    y = period["year"]
    m = period["month"]

    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    overall = compute_overall_total(db, merchant["id"], y, m)
    if overall["submission_status"] == "submitted":
        return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)

    if day < 1 or day > days_in_month(y, m):
        return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)

    visits = get_visits_for_month(db, merchant["id"], point_code, y, m)
    day_visits = visits.get(day, set())

    wd = weekday_of(y, m, day)
    is_fri_or_sat = wd in (4, 5)

    day_btn_text = "Ð£Ð±ÑÐ°ÑÑ Ð²ÑÑÐ¾Ð´" if "DAY" in day_visits else "ÐÐ¾Ð±Ð°Ð²Ð¸ÑÑ Ð²ÑÑÐ¾Ð´"
    inv_btn_text = "Ð£Ð±ÑÐ°ÑÑ Ð¿Ð¾Ð»Ð½ÑÐ¹ Ð¸Ð½Ð²ÐµÐ½Ñ" if "FULL_INVENT" in day_visits else "ÐÐ¾Ð±Ð°Ð²Ð¸ÑÑ Ð¿Ð¾Ð»Ð½ÑÐ¹ Ð¸Ð½Ð²ÐµÐ½Ñ"

    if not is_fri_or_sat:
        return RedirectResponse(url=f"/toggle-day?fio={escape(fio)}&point_code={escape(point_code)}&day={day}", status_code=303)

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>ÐÐµÐ¹ÑÑÐ²Ð¸Ðµ Ð¿Ð¾ Ð´Ð½Ñ</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <div class="brand">ÐÐºÑÑÐÐ¸Ð»Ð»</div>
            <h1>ÐÑÐ±Ð¾Ñ Ð´ÐµÐ¹ÑÑÐ²Ð¸Ñ</h1>
            <div class="subtitle">
                Ð¢Ð¾ÑÐºÐ°: {escape(point_code)}<br>
                ÐÐ°ÑÐ°: {day:02d}.{m:02d}.{y}
            </div>

            <a class="btn btn-small" href="/toggle-day?fio={escape(fio)}&point_code={escape(point_code)}&day={day}">
                {day_btn_text}
            </a>

            <a class="btn btn-secondary btn-small" href="/toggle-inventory?fio={escape(fio)}&point_code={escape(point_code)}&day={day}">
                {inv_btn_text}
            </a>

            <a class="back" href="/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}">â ÐÐ°Ð·Ð°Ð´ Ðº ÐºÐ°Ð»ÐµÐ½Ð´Ð°ÑÑ</a>
        </div>
    </div>
</body>
</html>
"""


@app.get("/toggle-day")
def toggle_day(
    fio: str,
    point_code: str,
    day: int,
    db: Session = Depends(get_db)
):
    period = get_active_period()
    y = period["year"]
    m = period["month"]

    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    overall = compute_overall_total(db, merchant["id"], y, m)
    if overall["submission_status"] == "submitted":
        return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)

    if 1 <= day <= days_in_month(y, m):
        toggle_day_visit(db, merchant["id"], point_code, y, m, day)

    return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)


@app.get("/toggle-inventory")
def toggle_inventory(
    fio: str,
    point_code: str,
    day: int,
    db: Session = Depends(get_db)
):
    period = get_active_period()
    y = period["year"]
    m = period["month"]

    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    overall = compute_overall_total(db, merchant["id"], y, m)
    if overall["submission_status"] == "submitted":
        return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)

    if 1 <= day <= days_in_month(y, m):
        wd = weekday_of(y, m, day)
        if wd in (4, 5):
            toggle_inventory_visit(db, merchant["id"], point_code, y, m, day)

    return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)


@app.get("/summary-page", response_class=HTMLResponse)
def summary_page(fio: str = "", db: Session = Depends(get_db)):
    period = get_active_period()
    merchant = get_merchant_by_fio(db, fio)
    overall = {"total": 0, "per_point": {}, "per_point_details": {}}

    if merchant:
        overall = compute_overall_total(db, merchant["id"], period["year"], period["month"])

    details_html = ""
    if overall["per_point_details"]:
        for point_code, d in overall["per_point_details"].items():
            details_html += f"""
            <details class="point-detail">
                <summary>{escape(point_code)} â {d["total"]} â½</summary>
                <div class="summary-content">
                    <div>Ð¡ Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¾Ð¹: {d["cnt_supply"]} Ã {d["rate_supply"]} â½ = {d["sum_supply"]} â½</div>
                    <div>ÐÐµÐ· Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¸: {d["cnt_no_supply"]} Ã {d["rate_no_supply"]} â½ = {d["sum_no_supply"]} â½</div>
                    <div>ÐÐ¾Ð»Ð½ÑÐ¹ Ð¸Ð½Ð²ÐµÐ½Ñ: {d["cnt_full_inv"]} Ã {d["rate_inventory"]} â½ = {d["sum_inventory"]} â½</div>
                    <div>ÐÐ¾ÑÐµÐ¼Ð°ÑÐ¸Ð½Ð°: {d["coffee_cnt"]} Ã {d["coffee_rate"]} â½ = {d["coffee_sum"]} â½</div>
                    <div><strong>ÐÑÐ¾Ð³Ð¾ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ: {d["total"]} â½</strong></div>
                </div>
            </details>
            """
    else:
        details_html = "<div class='hint' style='margin-top:10px'>ÐÐ¾ÐºÐ° Ð½ÐµÑ Ð¾ÑÐ¼ÐµÑÐµÐ½Ð½ÑÑ ÑÐ¾ÑÐµÐº Ð·Ð° ÑÑÐ¾Ñ Ð¼ÐµÑÑÑ.</div>"


    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>ÐÐ¾Ñ ÑÑÐ¼Ð¼Ð°</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <div class="brand">ÐÐºÑÑÐÐ¸Ð»Ð»</div>
            <h1>ÐÐ¾Ñ ÑÑÐ¼Ð¼Ð°</h1>
            <div class="subtitle">{escape(fio)}</div>

            <div class="sum-card">
                <div class="sum-title">ÐÐ±ÑÐ°Ñ ÑÑÐ¼Ð¼Ð° Ð·Ð° Ð¼ÐµÑÑÑ</div>
                <div class="sum-value">{overall["total"]} â½</div>
            </div>

            {details_html}

            <div class="hint">Ð¡ÐµÐ¹ÑÐ°Ñ Ð¾ÑÐºÑÑÑ Ð¿ÐµÑÐ¸Ð¾Ð´ Ð·Ð° {month_title(period["year"], period["month"])}.</div>

            <a class="back" href="/menu-page?fio={escape(fio)}">â ÐÐ°Ð·Ð°Ð´</a>
        </div>
    </div>
</body>
</html>
"""


@app.get("/admin-login", response_class=HTMLResponse)
def admin_login_page(error: str = ""):
    error_box = ""
    if error == "1":
        error_box = "<div class='error-box'>ÐÐµÐ²ÐµÑÐ½ÑÐ¹ Ð»Ð¾Ð³Ð¸Ð½ Ð¸Ð»Ð¸ Ð¿Ð°ÑÐ¾Ð»Ñ.</div>"

    env_box = ""
    if not ADMIN_LOGIN or not ADMIN_PASSWORD:
        env_box = "<div class='error-box'>Ð Render Ð½ÑÐ¶Ð½Ð¾ Ð·Ð°Ð´Ð°ÑÑ ADMIN_LOGIN Ð¸ ADMIN_PASSWORD.</div>"

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>ÐÐ´Ð¼Ð¸Ð½-Ð²ÑÐ¾Ð´</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <div class="brand">ÐÐºÑÑÐÐ¸Ð»Ð»</div>
            <h1>ÐÐ´Ð¼Ð¸Ð½ÐºÐ°</h1>
            <div class="subtitle">ÐÑÐ¾Ð´ Ð² Ð¾ÑÑÑÑ Ð¿Ð¾ ÑÐ²ÐµÑÐºÐ°Ð¼</div>

            {env_box}
            {error_box}

            <form method="post" action="/admin-login">
                <label for="login">ÐÐ¾Ð³Ð¸Ð½</label>
                <input id="login" name="login" type="text" required />

                <label for="password">ÐÐ°ÑÐ¾Ð»Ñ</label>
                <input id="password" name="password" type="password" required />

                <button class="btn" type="submit">ÐÐ¾Ð¹ÑÐ¸</button>
            </form>
        </div>
    </div>
</body>
</html>
"""


@app.post("/admin-login")
def admin_login_submit(login: str = Form(...), password: str = Form(...)):
    if not ADMIN_LOGIN or not ADMIN_PASSWORD:
        return RedirectResponse(url="/admin-login?error=1", status_code=303)

    if login != ADMIN_LOGIN or password != ADMIN_PASSWORD:
        return RedirectResponse(url="/admin-login?error=1", status_code=303)

    response = RedirectResponse(url="/admin-report", status_code=303)
    response.set_cookie(
        key="admin_auth",
        value=get_admin_cookie_value(),
        httponly=True,
        samesite="lax",
        secure=False,
        max_age=60 * 60 * 12,
    )
    return response


@app.get("/admin-logout")
def admin_logout():
    response = RedirectResponse(url="/admin-login", status_code=303)
    response.delete_cookie("admin_auth")
    return response


@app.get("/admin-report", response_class=HTMLResponse)
def admin_report(
    year: int | None = None,
    month: int | None = None,
    tu: str = "",
    status: str = "",
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    period = get_active_period()
    year = year or period["year"]
    month = month or period["month"]

    tu_filter = tu.strip() or None
    status_filter = status.strip() or None

    rows = get_admin_report_rows(db, year, month, tu_filter, status_filter)
    tu_values = get_all_tu_values(db)

    tu_options = "<option value=''>ÐÑÐµ Ð¢Ð£</option>"
    for item in tu_values:
        selected = "selected" if item == tu else ""
        tu_options += f"<option value='{escape(item)}' {selected}>{escape(item)}</option>"

    status_options = f"""
        <option value='' {'selected' if not status else ''}>ÐÑÐµ ÑÑÐ°ÑÑÑÑ</option>
        <option value='Ð½Ðµ Ð¾ÑÐ¿ÑÐ°Ð²Ð»ÐµÐ½Ð¾' {'selected' if status == 'Ð½Ðµ Ð¾ÑÐ¿ÑÐ°Ð²Ð»ÐµÐ½Ð¾' else ''}>ÐÐµ Ð¾ÑÐ¿ÑÐ°Ð²Ð»ÐµÐ½Ð¾</option>
        <option value='draft' {'selected' if status == 'draft' else ''}>Ð§ÐµÑÐ½Ð¾Ð²Ð¸Ðº</option>
        <option value='submitted' {'selected' if status == 'submitted' else ''}>ÐÑÐ¿ÑÐ°Ð²Ð»ÐµÐ½Ð¾</option>
    """

    rows_html = ""
    if rows:
        for r in rows:
            receipt_html = "â"
            if r["receipt_path"]:
                receipt_html = f"<a href='/{r['receipt_path']}' target='_blank'>ÐÑÐºÑÑÑÑ</a>"

            rows_html += f"""
            <tr>
                <td>{escape(r["fio"])}</td>
                <td>{escape(r["tu"]) if r["tu"] else "â"}</td>
                <td>{escape(r["point_code"])}</td>
                <td>{month_title(year, month)}</td>
                <td>{r["cnt_supply"]} / {r["sum_supply"]} â½</td>
                <td>{r["cnt_no_supply"]} / {r["sum_no_supply"]} â½</td>
                <td>{r["cnt_full_inv"]} / {r["sum_inventory"]} â½</td>
                <td>{r["coffee_cnt"]} Ã {r["coffee_rate"]} = {r["coffee_sum"]} â½</td>
                <td>{r["note_amount"]} â½<br>{escape(r["note_comment"]) if r["note_comment"] else "â"}</td>
                <td>{r["reimb_amount"]} â½<br>{escape(r["reimb_comment"]) if r["reimb_comment"] else "â"}</td>
                <td><strong>{r["point_total"]} â½</strong></td>
                <td>{escape(r["status"])}</td>
                <td>{f"<a href='/{r['reimb_receipt']}' target='_blank'>ÐÑÐºÑÑÑÑ</a>" if r["reimb_receipt"] else "â"}</td>
                <td>{escape(r["comment"]) if r["comment"] else "â"}</td>
            </tr>
            """
    else:
        rows_html = """
        <tr>
            <td colspan="13">ÐÐ¾ Ð²ÑÐ±ÑÐ°Ð½Ð½ÑÐ¼ ÑÐ¸Ð»ÑÑÑÐ°Ð¼ Ð´Ð°Ð½Ð½ÑÑ Ð½ÐµÑ.</td>
        </tr>
        """

    month_options = ""
    for m in range(1, 13):
        selected = "selected" if m == month else ""
        month_options += f"<option value='{m}' {selected}>{m:02d}</option>"

    export_query = f"year={year}&month={month}&tu={escape(tu)}&status={escape(status)}"

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>ÐÐ´Ð¼Ð¸Ð½-Ð¾ÑÑÑÑ</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card-wide">
            <div class="admin-actions">
                <div>
                    <div class="brand">ÐÐºÑÑÐÐ¸Ð»Ð»</div>
                    <h1>ÐÑÑÑÑ Ð¿Ð¾ ÑÐ²ÐµÑÐºÐ°Ð¼</h1>
                    <div class="subtitle">ÐÐ´Ð¼Ð¸Ð½-Ð¿Ð°Ð½ÐµÐ»Ñ</div>
                </div>
                <div class="admin-export-buttons">
                    <a class="btn btn-secondary btn-inline" href="/admin-data">Ð£Ð¿ÑÐ°Ð²Ð»ÐµÐ½Ð¸Ðµ Ð´Ð°Ð½Ð½ÑÐ¼Ð¸</a>
                    <a class="btn btn-secondary btn-inline" href="/admin-logout">ÐÑÐ¹ÑÐ¸</a>
                </div>
            </div>

            <form method="get" action="/admin-report">
                <div class="filter-grid">
                    <div>
                        <label for="year">ÐÐ¾Ð´</label>
                        <input id="year" name="year" type="number" value="{year}" />
                    </div>

                    <div>
                        <label for="month">ÐÐµÑÑÑ</label>
                        <select id="month" name="month">
                            {month_options}
                        </select>
                    </div>

                    <div>
                        <label for="tu">Ð¢ÐµÑÑÐ¸ÑÐ¾ÑÐ¸Ð°Ð»ÑÐ½ÑÐ¹ ÑÐ¿ÑÐ°Ð²Ð»ÑÑÑÐ¸Ð¹</label>
                        <select id="tu" name="tu">
                            {tu_options}
                        </select>
                    </div>

                    <div>
                        <label for="status">Ð¡ÑÐ°ÑÑÑ ÑÐ²ÐµÑÐºÐ¸</label>
                        <select id="status" name="status">
                            {status_options}
                        </select>
                    </div>
                </div>

                <button class="btn btn-inline" type="submit">ÐÑÐ¸Ð¼ÐµÐ½Ð¸ÑÑ ÑÐ¸Ð»ÑÑÑ</button>
            </form>

            <div class="admin-export-buttons">
                <a class="btn btn-secondary btn-inline" href="/admin-export-check?{export_query}">ÐÑÐ³ÑÑÐ·ÐºÐ° Ð´Ð»Ñ Ð¿ÑÐ¾Ð²ÐµÑÐºÐ¸</a>
                <a class="btn btn-secondary btn-inline" href="/admin-export-payroll?{export_query}">ÐÑÐ³ÑÑÐ·ÐºÐ° Ð² Ð²ÐµÐ´Ð¾Ð¼Ð¾ÑÑÑ</a>
                <a class="btn btn-secondary btn-inline" href="/admin-export-overlaps?{export_query}">ÐÑÐ³ÑÑÐ·ÐºÐ° Ð¿ÐµÑÐµÑÐµÑÐµÐ½Ð¸Ð¹</a>
            </div>

            <div class="hint">
                ÐÐµÑÐ¸Ð¾Ð´ Ð¾ÑÑÑÑÐ°: {month_title(year, month)}. ÐÑÐµÐ³Ð¾ ÑÑÑÐ¾Ðº: {len(rows)}.
            </div>

            <div class="table-wrap" style="margin-top:16px;">
                <table>
                    <thead>
                        <tr>
                            <th>Ð¤ÐÐ</th>
                            <th>Ð¢Ð£</th>
                            <th>Ð¢Ð¾ÑÐºÐ°</th>
                            <th>ÐÐµÑÑÑ</th>
                            <th>Ð¡ Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¾Ð¹</th>
                            <th>ÐÐµÐ· Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¸</th>
                            <th>ÐÐ½Ð²ÐµÐ½ÑÑ</th>
                            <th>ÐÐ¾ÑÐµÐ¼Ð°ÑÐ¸Ð½Ð°</th>
                            <th>ÐÑÐ¸Ð¼ÐµÑÐ°Ð½Ð¸Ðµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ</th>
                            <th>ÐÐ¾Ð·Ð¼ÐµÑÐµÐ½Ð¸Ðµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ</th>
                            <th>ÐÑÐ¾Ð³ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ</th>
                            <th>Ð¡ÑÐ°ÑÑÑ</th>
                            <th>Ð§ÐµÐº Ð¿Ð¾ ÑÐ¾ÑÐºÐµ</th>
                            <th>ÐÐ¾Ð¼Ð¼ÐµÐ½ÑÐ°ÑÐ¸Ð¹ Ð¼ÐµÑÑÑÐ°</th>
                        </tr>
                    </thead>
                    <tbody>
                        {rows_html}
                    </tbody>
                </table>
            </div>
        </div>
    </div>
</body>
</html>
"""


@app.get("/admin-data", response_class=HTMLResponse)
def admin_data_page(
    success: str = "",
    error: str = "",
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    period = get_active_period()
    tu_values = get_all_tu_values(db)

    tu_options = ""
    for item in tu_values:
        tu_options += f"<option value='{escape(item)}'>{escape(item)}</option>"

    info_box = ""
    if success:
        info_box += f"<div class='success-box'>{escape(success)}</div>"
    if error:
        info_box += f"<div class='error-box'>{escape(error)}</div>"

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Ð£Ð¿ÑÐ°Ð²Ð»ÐµÐ½Ð¸Ðµ Ð´Ð°Ð½Ð½ÑÐ¼Ð¸</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card-wide">
            <div class="admin-actions">
                <div>
                    <div class="brand">ÐÐºÑÑÐÐ¸Ð»Ð»</div>
                    <h1>Ð£Ð¿ÑÐ°Ð²Ð»ÐµÐ½Ð¸Ðµ Ð´Ð°Ð½Ð½ÑÐ¼Ð¸</h1>
                    <div class="subtitle">ÐÐ°Ð³ÑÑÐ·ÐºÐ° ÑÐ°Ð¹Ð»Ð¾Ð² Ð¸ Ð¾ÑÐ¸ÑÑÐºÐ° Ð¼ÐµÑÑÑÐ°</div>
                </div>
                <div class="admin-export-buttons">
                    <a class="btn btn-secondary btn-inline" href="/admin-report">ÐÐ°Ð·Ð°Ð´ Ðº Ð¾ÑÑÑÑÑ</a>
                    <a class="btn btn-secondary btn-inline" href="/admin-logout">ÐÑÐ¹ÑÐ¸</a>
                </div>
            </div>

            {info_box}

            <div class="data-grid">
                <div class="detail-card">
                    <div class="detail-title">ÐÐ°Ð³ÑÑÐ·ÐºÐ° Ð¿Ð¾ÑÑÐ°Ð²Ð¾Ðº</div>
                    <form method="post" action="/admin-upload-supplies" enctype="multipart/form-data">
                        <label for="supplies_file">Ð¤Ð°Ð¹Ð» Ð¿Ð¾ÑÑÐ°Ð²Ð¾Ðº</label>
                        <input id="supplies_file" name="file" type="file" accept=".xlsx" required />
                        <button class="btn" type="submit">ÐÐ°Ð³ÑÑÐ·Ð¸ÑÑ Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¸</button>
                    </form>
                </div>

                <div class="detail-card">
                    <div class="detail-title">ÐÐ°Ð³ÑÑÐ·ÐºÐ° ÑÑÐ°Ð²Ð¾Ðº</div>
                    <form method="post" action="/admin-upload-rates" enctype="multipart/form-data">
                        <label for="rates_year">ÐÐ¾Ð´</label>
                        <input id="rates_year" name="year" type="number" value="{period["year"]}" required />

                        <label for="rates_month">ÐÐµÑÑÑ</label>
                        <input id="rates_month" name="month" type="number" value="{period["month"]}" min="1" max="12" required />

                        <label for="rates_file">Ð¤Ð°Ð¹Ð» ÑÑÐ°Ð²Ð¾Ðº</label>
                        <input id="rates_file" name="file" type="file" accept=".xlsx" required />

                        <button class="btn" type="submit">ÐÐ°Ð³ÑÑÐ·Ð¸ÑÑ ÑÑÐ°Ð²ÐºÐ¸</button>
                    </form>
                </div>

                <div class="detail-card">
                    <div class="detail-title">ÐÐ°Ð³ÑÑÐ·ÐºÐ° Ð¼ÐµÑÑÐµÐ¹</div>
                    <form method="post" action="/admin-upload-merchants" enctype="multipart/form-data">
                        <label for="merchants_tu">Ð¢ÐµÑÑÐ¸ÑÐ¾ÑÐ¸Ð°Ð»ÑÐ½ÑÐ¹ ÑÐ¿ÑÐ°Ð²Ð»ÑÑÑÐ¸Ð¹</label>
                        <input id="merchants_tu" name="tu" type="text" placeholder="ÐÐ°Ð¿ÑÐ¸Ð¼ÐµÑ: Ð¥ÑÑÐ¿Ð¾Ð²" required />

                        <label for="merchants_file">Ð¤Ð°Ð¹Ð» Ð¼ÐµÑÑÐµÐ¹</label>
                        <input id="merchants_file" name="file" type="file" accept=".xlsx" required />

                        <button class="btn" type="submit">ÐÐ°Ð³ÑÑÐ·Ð¸ÑÑ Ð¼ÐµÑÑÐµÐ¹</button>
                    </form>
                </div>

                <div class="detail-card">
                    <div class="detail-title">ÐÑÐ¸ÑÑÐºÐ° Ð¼ÐµÑÑÑÐ°</div>
                    <form method="post" action="/admin-clear-month">
                        <label for="clear_year">ÐÐ¾Ð´</label>
                        <input id="clear_year" name="year" type="number" value="{period["year"]}" required />

                        <label for="clear_month">ÐÐµÑÑÑ</label>
                        <input id="clear_month" name="month" type="number" value="{period["month"]}" min="1" max="12" required />

                        <button class="btn btn-danger" type="submit">ÐÑÐ¸ÑÑÐ¸ÑÑ Ð´Ð°Ð½Ð½ÑÐµ Ð¼ÐµÑÑÑÐ°</button>
                    </form>
                </div>

                <div class="detail-card">
                    <div class="detail-title">ÐÑÐ¸ÑÑÐºÐ° Ð¼ÐµÑÑÐµÐ¹ Ð¿Ð¾ Ð¢Ð£</div>
                    <form method="post" action="/admin-clear-merchants">
                        <label for="clear_tu">Ð¢ÐµÑÑÐ¸ÑÐ¾ÑÐ¸Ð°Ð»ÑÐ½ÑÐ¹ ÑÐ¿ÑÐ°Ð²Ð»ÑÑÑÐ¸Ð¹</label>
                        <select id="clear_tu" name="tu" required>
                            {tu_options}
                        </select>

                        <button class="btn btn-danger" type="submit">Ð£Ð´Ð°Ð»Ð¸ÑÑ Ð¼ÐµÑÑÐµÐ¹ ÑÑÐ¾Ð³Ð¾ Ð¢Ð£</button>
                    </form>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
"""


@app.post("/admin-upload-supplies")
async def admin_upload_supplies(
    file: UploadFile = File(...),
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    try:
        result = import_supplies_xlsx(db, file.file)
        msg = f"ÐÐ¾ÑÑÐ°Ð²ÐºÐ¸ Ð·Ð°Ð³ÑÑÐ¶ÐµÐ½Ñ: ÑÑÑÐ¾Ðº {result['loaded_rows']}, ÑÐ¾ÑÐµÐº {result['loaded_points']}."
        return RedirectResponse(url=f"/admin-data?success={msg}", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/admin-data?error={str(e)}", status_code=303)


@app.post("/admin-upload-rates")
async def admin_upload_rates(
    year: int = Form(...),
    month: int = Form(...),
    file: UploadFile = File(...),
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    try:
        result = import_rates_xlsx(db, file.file, year, month)
        msg = f"Ð¡ÑÐ°Ð²ÐºÐ¸ Ð·Ð°Ð³ÑÑÐ¶ÐµÐ½Ñ: ÑÑÑÐ¾Ðº {result['loaded_rows']}."
        return RedirectResponse(url=f"/admin-data?success={msg}", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/admin-data?error={str(e)}", status_code=303)


@app.post("/admin-upload-merchants")
async def admin_upload_merchants(
    tu: str = Form(...),
    file: UploadFile = File(...),
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    try:
        result = import_merchants_xlsx(db, file.file, tu)
        msg = f"ÐÐµÑÑÐ¸ Ð·Ð°Ð³ÑÑÐ¶ÐµÐ½Ñ: ÑÑÑÐ¾Ðº {result['loaded_rows']}."
        return RedirectResponse(url=f"/admin-data?success={msg}", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/admin-data?error={str(e)}", status_code=303)


@app.post("/admin-clear-month")
def admin_clear_month(
    year: int = Form(...),
    month: int = Form(...),
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    result = clear_month_data(db, year, month)
    msg = (
        f"ÐÐµÑÑÑ Ð¾ÑÐ¸ÑÐµÐ½. ÐÐ¸Ð·Ð¸ÑÑ: {result['deleted_visits']}, "
        f"Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¸: {result['deleted_supplies']}, ÑÑÐ°Ð²ÐºÐ¸: {result['deleted_rates']}, "
        f"Ð¼ÐµÑÑÑÐ½ÑÐµ ÑÐ²ÐµÑÐºÐ¸: {result['deleted_monthly']}, ÐºÐ¾ÑÑÐµÐºÑÐ¸ÑÐ¾Ð²ÐºÐ¸ Ð¿Ð¾ ÑÐ¾ÑÐºÐ°Ð¼: {result.get('deleted_point_adjustments', 0)}."
    )
    return RedirectResponse(url=f"/admin-data?success={msg}", status_code=303)


@app.post("/admin-clear-merchants")
def admin_clear_merchants(
    tu: str = Form(...),
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    deleted = clear_merchants_by_tu(db, tu)
    msg = f"Ð£Ð´Ð°Ð»ÐµÐ½Ð¾ Ð¼ÐµÑÑÐµÐ¹ Ð¢Ð£ {tu}: {deleted}."
    return RedirectResponse(url=f"/admin-data?success={msg}", status_code=303)


@app.get("/admin-export-check")
def admin_export_check(
    year: int,
    month: int,
    tu: str = "",
    status: str = "",
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    rows = get_admin_report_rows(
        db=db,
        y=year,
        m=month,
        tu=tu.strip() or None,
        status=status.strip() or None
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "ÐÑÐ¾Ð²ÐµÑÐºÐ°"

    ws.append([
        "Ð¤ÐÐ",
        "Ð¢Ð£",
        "Ð¢Ð¾ÑÐºÐ°",
        "ÐÐµÑÑÑ",
        "ÐÑÑÐ¾Ð´Ñ Ñ Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¾Ð¹ (ÐºÐ¾Ð»-Ð²Ð¾)",
        "ÐÑÑÐ¾Ð´Ñ Ñ Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¾Ð¹ (ÑÑÐ¼Ð¼Ð°)",
        "ÐÑÑÐ¾Ð´Ñ Ð±ÐµÐ· Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¸ (ÐºÐ¾Ð»-Ð²Ð¾)",
        "ÐÑÑÐ¾Ð´Ñ Ð±ÐµÐ· Ð¿Ð¾ÑÑÐ°Ð²ÐºÐ¸ (ÑÑÐ¼Ð¼Ð°)",
        "ÐÐ¾Ð»Ð½ÑÐµ Ð¸Ð½Ð²ÐµÐ½ÑÑ (ÐºÐ¾Ð»-Ð²Ð¾)",
        "ÐÐ¾Ð»Ð½ÑÐµ Ð¸Ð½Ð²ÐµÐ½ÑÑ (ÑÑÐ¼Ð¼Ð°)",
        "ÐÐ¾ÑÐµÐ¼Ð°ÑÐ¸Ð½Ð° (ÐºÐ¾Ð»-Ð²Ð¾)",
        "ÐÐ¾ÑÐµÐ¼Ð°ÑÐ¸Ð½Ð° (ÑÑÐ¼Ð¼Ð°)",
        "ÐÑÐ¸Ð¼ÐµÑÐ°Ð½Ð¸Ðµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ (ÑÑÐ¼Ð¼Ð°)",
        "ÐÑÐ¸Ð¼ÐµÑÐ°Ð½Ð¸Ðµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ (ÐºÐ¾Ð¼Ð¼ÐµÐ½ÑÐ°ÑÐ¸Ð¹)",
        "ÐÐ¾Ð·Ð¼ÐµÑÐµÐ½Ð¸Ðµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ (ÑÑÐ¼Ð¼Ð°)",
        "ÐÐ¾Ð·Ð¼ÐµÑÐµÐ½Ð¸Ðµ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ (ÐºÐ¾Ð¼Ð¼ÐµÐ½ÑÐ°ÑÐ¸Ð¹)",
        "Ð§ÐµÐº Ð¿Ð¾ ÑÐ¾ÑÐºÐµ",
        "Ð¡ÑÐ°ÑÑÑ",
        "ÐÐ¾Ð¼Ð¼ÐµÐ½ÑÐ°ÑÐ¸Ð¹ Ð¼ÐµÑÑÑÐ°",
        "ÐÑÐ¾Ð³ Ð¿Ð¾ ÑÐ¾ÑÐºÐµ"
    ])

    for r in rows:
        ws.append([
            r["fio"],
            r["tu"],
            r["point_code"],
            month_title(year, month),
            r["cnt_supply"],
            r["sum_supply"],
            r["cnt_no_supply"],
            r["sum_no_supply"],
            r["cnt_full_inv"],
            r["sum_inventory"],
            r["coffee_cnt"],
            r["coffee_sum"],
            r["note_amount"],
            r["note_comment"],
            r["reimb_amount"],
            r["reimb_comment"],
            r["reimb_receipt"] or "",
            r["status"],
            r["comment"],
            r["point_total"]
        ])

    style_sheet(ws)
    return build_excel_response(wb, f"proverka_{year}_{month:02d}.xlsx")


@app.get("/admin-export-payroll")
def admin_export_payroll(
    year: int,
    month: int,
    tu: str = "",
    status: str = "",
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    rows = get_admin_payroll_rows(
        db=db,
        y=year,
        m=month,
        tu=tu.strip() or None,
        status=status.strip() or None
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "ÐÐµÐ´Ð¾Ð¼Ð¾ÑÑÑ"

    ws.append([
        "Ð¤ÐÐ",
        "Ð¢Ð£",
        "Ð¡ÑÐ¼Ð¼Ð° Ð¿Ð¾ Ð¼ÐµÑÑÑ",
        "Ð¡ÑÐ¼Ð¼Ð° Ð² Ð²ÐµÐ´Ð¾Ð¼Ð¾ÑÑÑ (/0.87, Ð¾ÐºÑÑÐ³Ð»ÐµÐ½Ð¸Ðµ Ð²Ð²ÐµÑÑ)",
        "Ð¡ÑÐ°ÑÑÑ"
    ])

    for r in rows:
        ws.append([
            r["fio"],
            r["tu"],
            r["clean_total"],
            r["payroll_total"],
            r["status"]
        ])

    style_sheet(ws)
    return build_excel_response(wb, f"vedomost_{year}_{month:02d}.xlsx")


@app.get("/admin-export-overlaps")
def admin_export_overlaps(
    year: int,
    month: int,
    tu: str = "",
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    rows = get_intersections_rows(
        db=db,
        y=year,
        m=month,
        tu=tu.strip() or None
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "ÐÐµÑÐµÑÐµÑÐµÐ½Ð¸Ñ"

    ws.append([
        "ÐÐ°ÑÐ°",
        "Ð¢Ð¾ÑÐºÐ°",
        "ÐÐµÑÑ 1",
        "Ð¢Ð£ 1",
        "Ð¡Ð»Ð¾Ñ 1",
        "ÐÐµÑÑ 2",
        "Ð¢Ð£ 2",
        "Ð¡Ð»Ð¾Ñ 2"
    ])

    for r in rows:
        ws.append([
            r["visit_date"],
            r["point_code"],
            r["fio1"],
            r["tu1"],
            r["slot1"],
            r["fio2"],
            r["tu2"],
            r["slot2"],
        ])

    style_sheet(ws)
    return build_excel_response(wb, f"peresecheniya_{year}_{month:02d}.xlsx")
